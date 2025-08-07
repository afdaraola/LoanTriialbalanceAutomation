import sys
import math
import os
import openpyxl 
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font,PatternFill,Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import pymysql
#import pandas as pd
import mysql.connector
import re


#! python3

print('start calling up the excel')



#set the directory 
os.chdir("C:\\Users\\daram\\pythonscripts")

#display working directory
print(os.getcwd())

#extract parameters from property file 
seperator="=="
keys ={}

with open('config.txt') as f:
    for line in f:
        if seperator in line:
            name, value = line.split(seperator,1)
            keys[name.strip()] = value.strip()

print("parameters -> ", keys)


#load the workbook 
wb = openpyxl.load_workbook("Template.xlsx")

 # Center-align the content of the cell
alignment = Alignment(horizontal='center', vertical='center')

#greyout some rows 
# Define the grey fill style
grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

grey_fill_excptn_area = PatternFill(start_color="d4ebf2", end_color="d4ebf2", fill_type="solid")

#define default value
def defaultvalue(x):
    return 0 if x is None else x

#get the acctive sheet 
sheetname = wb["Cover Sheet"] 


# Define border styles
top_border = Border(
    top=Side(style='medium'),
    bottom=Side(style='thin')
    )

right_border = Border( 
    right=Side(style='medium')
    )

bottom_border = Border(
    top=Side(style='thin'),
    bottom=Side(style='medium')
    )

# Define border styles
thick_border = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

# Define thin border styles
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
    )

#display working directory
print(os.getcwd())


formattingAmount = keys['formatt']
#define custant 
NETVAR= keys['variance']


# Create a comment
print("Add comment of accrued through date >>>>>")
comment_text = keys['commentText'] 
author = keys['authorComment']
comment = Comment(comment_text,author)

# Add the comment to the cell
sheetname['H4'].comment = comment



print("Number format parttern --> "+formattingAmount)

#set the headers 
#merge cells 
sheetname.merge_cells("A2:H2")
 #set title 
sheetname['A2'] = "Loan Conversion balance Sheet" 

# Define the range of cells to apply borders
cell_range = sheetname['A2:H2']

# Apply the border to each cell in the range
for row in sheetname['A7:H7']:
    for cell in row:
        cell.border = thick_border
        cell.fill = grey_fill

#fill the title 
for row in cell_range:
    for cell in row:
        cell.fill = grey_fill

# Define the range of cells to apply borders 
sheet_range = sheetname['B3:B53']

# Apply the border to each cell in the range entire cell
for row in sheet_range:
   for cell in row:
        cell.border = right_border

for row in sheetname['D3:D53']:
   for cell in row:
        cell.border = right_border
        
for row in sheetname['F3:F53']:
   for cell in row:
        cell.border = right_border
        
for row in sheetname['H3:H53']:
   for cell in row:
        cell.border = right_border
        

# Center-align the content of the cell
sheetname['A2'].alignment = Alignment(horizontal='center', vertical='center')

# Apply bold formatting to the cell
bold_font = Font(bold=True)
sheetname['A2'].font = bold_font

 #set parameters 
sheetname['A3'] = "CLIENT NAME:"
sheetname['C3'] = keys['clientname']  
sheetname['E3'] = "CONVERSION/TAPE DATE:" 
sheetname['F3']=keys['conversiondate']
sheetname['E4'] = "ACCRUED THROUGH DATE:" 
sheetname['F4']=keys['accruedthrough']

 

sheetname['A5'] = "BALANCED BY:" 
sheetname['C5']=keys['balancedby']
sheetname['A6']= "MANAGER BRANCH:" 
sheetname['C6']=keys['branch']


sheetname['C6'].font = bold_font

sheetname['A7'] = "ITEM/TASK"  
sheetname['A7'].font = bold_font
sheetname['A7'].alignment = alignment
#Setting dmpdsts
#merge cells 
sheetname.merge_cells("C7:D7")
sheetname['C7'] = "CURRENT SYSTEM" 
sheetname['C7'].font = bold_font

sheetname['C7'].alignment = alignment


#Setting DNA
#merge cells 
sheetname.merge_cells("E7:F7")
sheetname['E7'] = "DNA SYSTEM" 
sheetname['E7'].font = bold_font
# Center-align the content of the cell
sheetname['E7'].alignment = alignment
 
 #Setting Variance
 #merge cells 
sheetname.merge_cells("G7:H7")
sheetname['G7'] = "VARIANCES- SEE EXPLANATIONS BELOW" 
sheetname['G7'].font = bold_font
# Center-align the content of the cell
sheetname['G7'].alignment = alignment
 
list_with_variance={}
 
sheetname['A8'] = "Principal Balance"
sheetname['E8'] = "CM_TRIAL"
sheetname['E9'] = "CN_TRIAL"
sheetname['E10'] = "MG_TRIAL"
sheetname['G8'] = "(1)"
sheetname['H8'] = "=D8+D9+D10-F8-F9-F10"
#format currency
sheetname['H8'].number_format = formattingAmount 

list_with_variance['PRINCIPALBALANCE'] = 'H8'

sheetname['A11'] = "Accrued Interest"
sheetname['E11'] = "CM_TRIAL"
sheetname['E12'] = "CN_TRIAL"
sheetname['E13'] = "MG_TRIAL"
sheetname['G11'] = "(2)"

sheetname['A14'] = "Non-Accrued Interest"
sheetname['E14'] = "CM_TRIAL"
sheetname['E15'] = "CN_TRIAL"
sheetname['E16'] = "MG_TRIAL"
sheetname['H11'] = "=D11+D12+D13-F11-F12-F13"
#format currency
sheetname['H11'].number_format = formattingAmount 

list_with_variance['ACCRUEDINTEREST'] = 'H11'

sheetname['A17'] = "Escrow Balance"
sheetname['E17'] = "CM_TRIAL"
sheetname['E18'] = "CN_TRIAL"
sheetname['E19'] = "MG_TRIAL"
sheetname['G17'] = "(3)"
sheetname['H17'] = "=D17+D18+D19-F17-F18-F19"
#format currency
sheetname['H17'].number_format = formattingAmount 

list_with_variance['ESCROWBALANCE'] = 'H17'



sheetname['A20'] = "Late Charge"
sheetname['E20'] = "CM_TRIAL"
sheetname['E21'] = "CN_TRIAL"
sheetname['E22'] = "MG_TRIAL"
sheetname['G20'] = "(4)"
sheetname['H20'] = "=D20+D21+D22-F20-F21-F22"

#format currency
sheetname['H20'].number_format = formattingAmount 

list_with_variance['LATECHARGES'] = 'H20'

sheetname['A23'] = "Defferred Fees"
sheetname['E23'] = "LN_AMRT" 
sheetname['G23'] = "(5)"
sheetname['H23'] = "=D23+D23+D25-F23"

#format currency
sheetname['H23'].number_format = formattingAmount 

list_with_variance['DEFFEREDFEES'] = 'H23'

sheetname['A26'] = "Defferred Costs"
sheetname['E26'] = "LN_AMRT" 
sheetname['G26'] = "(6)"
sheetname['H26'] = "=D26+D27+D28-F27"
#format currency
sheetname['H26'].number_format = formattingAmount 

list_with_variance['DEFFEREDCOSTS'] = 'H26'

sheetname['A29'] = "Original Comm (LOC)"
sheetname['E29'] = "CM_TRIAL"
sheetname['E30'] = "CN_TRIAL"
sheetname['E31'] = "MG_TRIAL"
sheetname['E32'] = "ML_TRIAL"
sheetname['G29'] = "(7)"
sheetname['H29'] = "=D29+D30+D31-F29-F30-F31-F32"

#format currency
sheetname['H29'].number_format = formattingAmount 

list_with_variance['CREDITLIMIT'] = 'H29'

sheetname['A33'] = "Charge-off Balance"
sheetname['E33'] = "CM_TRIAL"
sheetname['E34'] = "CN_TRIAL"
sheetname['E35'] = "MG_TRIAL"
sheetname['G33'] = "(8)"
sheetname['H33'] = "=D33+D34+D35-F33-F34-F35" 
#format currency
sheetname['H33'].number_format = formattingAmount 

list_with_variance['CHARGEOFFBALANCE'] = 'H33'


sheetname['A36'] = "Unapplied Balance"
sheetname['E36'] = "CM_TRIAL"
sheetname['E37'] = "CN_TRIAL"
sheetname['E38'] = "MG_TRIAL"
sheetname['G36'] = "(9)"
sheetname['H36'] = "=D36+D37+D38-F36-F37-F38" 
#format currency
sheetname['H36'].number_format = formattingAmount 

list_with_variance['UNAPPLIEDBALANCE'] = 'H36'

sheetname['A39'] = "Investor Balance"
sheetname['E39'] = "CM_TRIAL/IL_TRIAL"
sheetname['E40'] = "CN_TRIAL/IL_TRIAL"
sheetname['E41'] = "MG_TRIAL/IL_TRIAL"
sheetname['G39'] = "(10)"
sheetname['H39'] = "=D39+D40+D41-F39-F40-F41" 
#format currency
sheetname['H39'].number_format = formattingAmount 

list_with_variance['INVESTORBALANCE'] = 'H39'

sheetname['A42'] = "Participation"
sheetname['E42'] = "CM_TRIAL"
sheetname['E43'] = "CN_TRIAL"
sheetname['E44'] = "MG_TRIAL"
sheetname['G42'] = "(11)"
sheetname['H42'] = "=D42+D43+D44-F42-F43-F44" 

#format currency
sheetname['H42'].number_format = formattingAmount 

list_with_variance['PARTICIPATION'] = 'H42'

sheetname['A45'] = "YTD Interest"
sheetname['E45'] = "CM_TRIAL"
sheetname['E46'] = "CN_TRIAL"
sheetname['E47'] = "MG_TRIAL"
sheetname['G45'] = "(12)"
sheetname['H45'] = "=D45+D46+D47-F45-F46-F47" 

#format currency
sheetname['H45'].number_format = formattingAmount 

list_with_variance['YTDINTEREST'] = 'H45'

sheetname['A48'] = "YTD Late Charge Paid"
sheetname['E48'] = "MM_YTD"
sheetname['G48'] = "(13)"
sheetname['H48'] = "=D48-F48-F49-F50"

#format currency
sheetname['H48'].number_format = formattingAmount 

list_with_variance['YTDLATECHARGEPAID'] = 'H48'

sheetname['A51'] = "Shadow Accounting IPTP"
sheetname['E51'] = "CM_TRIAL"
sheetname['E52'] = "CN_TRIAL"
sheetname['E53'] = "MG_TRIAL"
sheetname['G51'] = "(14)"
sheetname['H51'] = "=D45+D46+D47-F45-F46-F47" 

#format currency
sheetname['H51'].number_format = formattingAmount 

list_with_variance['SHADOWACCOUNTINGIPTP'] = 'H51'

#Connection to DB 
print("Connection to DB.....")
try:
     # Establish the connection
    conn = pymysql.connect(host = keys['dbhost'],
    port = int(3306),
    user = keys['dbusername'],
    password = keys['dbpassword'],
    db = keys['dbname'])
except Exception as e:
     print(e)
# Check if the connection is successful
if conn:
     print ("connection successful") 

# Create a cursor object
cursor = conn.cursor()
 
# Write and execute the SELECT query
query = keys["loanbalancing"] #query in property file 
cursor.execute(query)

# Fetch all rows from the executed query
results = cursor.fetchall()

# Print the results
for row in results:
    #print(row)
    
    if row[1] =='PRINCIPALBALANCE':
        print("BAL CATEGORY ->>", row[1] , row[4])
        if row[4]=='CML':
            sheetname['C8'] = row[2]
            sheetname['D8'] = row[3]
            sheetname['F8'] = row[5]
        if row[4]=='CNS':
            sheetname['C9'] = row[2]
            sheetname['D9'] = row[3]
            sheetname['F9'] = row[5]
        if row[4]=='MTG':
            sheetname['C10'] = row[2]
            sheetname['D10'] = row[3]
            sheetname['F10'] = row[5]
    elif row[1] =='ACCRUEDINTEREST':
        print("BAL CATEGORY ->>", row[1] , row[4])
        if row[4]=='CML':
            sheetname['C11'] = row[2]
            sheetname['D11'] = row[3]
            sheetname['F11'] = row[5]
        if row[4]=='CNS':
            sheetname['C12'] = row[2]
            sheetname['D12'] = row[3]
            sheetname['F12'] = row[5]
        if row[4]=='MTG':
            sheetname['C13'] = row[2]
            sheetname['D13'] = row[3]
            sheetname['F13'] = row[5] 
    elif row[1] =='ESCROWBALANCE':
        print("BAL CATEGORY ->>", row[1] , row[4])
        if row[4]=='CML':
            sheetname['C17'] = row[2]
            sheetname['D17'] = row[3]
            sheetname['F17'] = row[5]
        if row[4]=='CNS':
            sheetname['C18'] = row[2]
            sheetname['D18'] = row[3]
            sheetname['F18'] = row[5]
        if row[4]=='MTG':
            sheetname['C19'] = row[2]
            sheetname['D19'] = row[3]
            sheetname['F19'] = row[5] 
    elif row[1] =='LATECHARGES':
        print("BAL CATEGORY ->>", row[1] , row[4])
        if row[4]=='CML':
            sheetname['C20'] = row[2]
            sheetname['D20'] = row[3]
            sheetname['F20'] = row[5]
        if row[4]=='CNS':
            sheetname['C21'] = row[2]
            sheetname['D21'] = row[3]
            sheetname['F21'] = row[5]
        if row[4]=='MTG':
            sheetname['C22'] = row[2]
            sheetname['D22'] = row[3]
            sheetname['F22'] = row[5] 
    elif row[1] =='DEFFEREDFEES':
        print("BAL CATEGORY ->>", row[1] , row[4])
        if row[4]=='CML':
            sheetname['C23'] = row[2]
            sheetname['D23'] = row[3]
            sheetname['F23'] = row[5]
        if row[4]=='CNS':
            sheetname['C24'] = row[2]
            sheetname['D24'] = row[3]
            sheetname['F24'] = row[5]
        if row[4]=='MTG':
            sheetname['C25'] = row[2]
            sheetname['D25'] = row[3]
            sheetname['F25'] = row[5] 
    elif row[1] =='DEFFEREDCOSTS':
        print("BAL CATEGORY ->>", row[1] , row[4])
        if row[4]=='CML':
            sheetname['C26'] = row[2]
            sheetname['D26'] = row[3]
            sheetname['F26'] = row[5]
        if row[4]=='CNS':
            sheetname['C27'] = row[2]
            sheetname['D27'] = row[3]
            sheetname['F27'] = row[5]
        if row[4]=='MTG':
            sheetname['C28'] = row[2]
            sheetname['D28'] = row[3]
            sheetname['F28'] = row[5] 
    elif row[1] =='CREDITLIMIT':
        print("BAL CATEGORY ->>", row[1] , row[4])
        if row[4]=='CML':
            sheetname['C29'] = row[2]
            sheetname['D29'] = row[3]
            sheetname['F29'] = row[5]
        if row[4]=='CNS':
            sheetname['C30'] = row[2]
            sheetname['D30'] = row[3]
            sheetname['F30'] = row[5]
        if row[4]=='MTG':
            sheetname['C31'] = row[2]
            sheetname['D31'] = row[3]
            sheetname['F31'] = row[5]
        if row[4]=='MLN':
            sheetname['C32'] = row[2]
            sheetname['D32'] = row[3]
            sheetname['F32'] = row[5] 
    elif row[1] =='CHARGEOFFBALANCE':
        print("BAL CATEGORY ->>", row[1] , row[4])
        if row[4]=='CML':
            sheetname['C33'] = row[2]
            sheetname['D33'] = row[3]
            sheetname['F33'] = row[5]
        if row[4]=='CNS':
            sheetname['C34'] = row[2]
            sheetname['D34'] = row[3]
            sheetname['F34'] = row[5]
        if row[4]=='MTG':
            sheetname['C35'] = row[2]
            sheetname['D35'] = row[3]
            sheetname['F35'] = row[5] 
    elif row[1] =='UNAPPLIEDBALANCE':
        print("BAL CATEGORY ->>", row[1] , row[4])
        if row[4]=='CML':
            sheetname['C36'] = row[2]
            sheetname['D36'] = row[3]
            sheetname['F36'] = row[5]
        if row[4]=='CNS':
            sheetname['C37'] = row[2]
            sheetname['D37'] = row[3]
            sheetname['F37'] = row[5]
        if row[4]=='MTG':
            sheetname['C38'] = row[2]
            sheetname['D38'] = row[3]
            sheetname['F38'] = row[5] 
    elif row[1] =='INVESTORBALANCE':
        print("BAL CATEGORY ->>", row[1] , row[4])
        if row[4]=='CML':
            sheetname['C39'] = row[2]
            sheetname['D39'] = row[3]
            sheetname['F39'] = row[5]
        if row[4]=='CNS':
            sheetname['C40'] = row[2]
            sheetname['D40'] = row[3]
            sheetname['F40'] = row[5]
        if row[4]=='MTG':
            sheetname['C41'] = row[2]
            sheetname['D41'] = row[3]
            sheetname['F41'] = row[5] 
    elif row[1] =='PARTICIPATION':
        print("BAL CATEGORY ->>", row[1] , row[4])
        if row[4]=='CML':
            sheetname['C42'] = row[2]
            sheetname['D42'] = row[3]
            sheetname['F42'] = row[5]
        if row[4]=='CNS':
            sheetname['C43'] = row[2]
            sheetname['D43'] = row[3]
            sheetname['F43'] = row[5]
        if row[4]=='MTG':
            sheetname['C44'] = row[2]
            sheetname['D44'] = row[3]
            sheetname['F44'] = row[5] 
    elif row[1] =='YTDINTEREST':
        print("BAL CATEGORY ->>", row[1] , row[4])
        if row[4]=='CML':
            sheetname['C45'] = row[2]
            sheetname['D45'] = row[3]
            sheetname['F45'] = row[5]
        if row[4]=='CNS':
            sheetname['C46'] = row[2]
            sheetname['D46'] = row[3]
            sheetname['F46'] = row[5]
        if row[4]=='MTG':
            sheetname['C47'] = row[2]
            sheetname['D47'] = row[3]
            sheetname['F47'] = row[5] 
    elif row[1] =='YTDLATECHARGEPAID':
        print("BAL CATEGORY ->>", row[1] , row[4])
        if row[4]=='CML':
            sheetname['C48'] = row[2] 
            sheetname['D48'] = row[3]
            sheetname['F48'] = row[5]
        if row[4]=='CNS':
            sheetname['C49'] = row[2] 
            sheetname['D49'] = row[3]
            sheetname['F49'] = row[5]
        if row[4]=='MTG':
            sheetname['C50'] = row[2] 
            sheetname['D50'] = row[3]
            sheetname['F50'] = row[5] 
    elif row[1] =='SHADOWACCOUNTINGIPTP':
        print("BAL CATEGORY ->>", row[1] , row[4])
        if row[4]=='CML':
            sheetname['C51'] = row[2]
            sheetname['D51'] = row[3]
            sheetname['F51'] = row[5]
        if row[4]=='CNS':
            sheetname['C52'] = row[2]
            sheetname['D52'] = row[3]
            sheetname['F52'] = row[5]
        if row[4]=='MTG':
            sheetname['C53'] = row[2]
            sheetname['D53'] = row[3]
            sheetname['F53'] = row[5] 

else:
    print ("Clossing connection")
# Close the connection when done
cursor.close()
conn.close()    


#bold row E
for row in range(8, 55):
    cell = sheetname[f"E{row}"]
    cell.font =bold_font
    cell.alignment = alignment
    #format currency amount DNA values
    currencycellDNA = sheetname[f"F{row}"] 
    currencycellDNA.number_format = formattingAmount 
    #format currency amount Dmpdata values
    currencycellDmp = sheetname[f"D{row}"] 
    currencycellDmp.number_format = formattingAmount 
    
    

# Add themed text to the column
color_font = Font(color="FF0000")

sheetname['A60'] = "Explanation of any out of balance conditions:" 


# Define the range of cells to apply borders
shadow_top_border_range = sheetname['A51:H51']

# Apply the border to top cell in the range
for row in shadow_top_border_range:
    for top_cell in row:
        top_cell.border = top_border

# Define the range of cells to apply borders
shadow_bottom_border_range = sheetname['A53:H53']

# Apply the border to top cell in the range
for row in shadow_bottom_border_range:
    for bottom_cell in row:
        bottom_cell.border = bottom_border
        
for row in sheetname['A8:H8']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border       
 
for row in sheetname['A10:H10']:
    for prin_bottom_cell in row:
        prin_bottom_cell.border = bottom_border  


for row in sheetname['A17:H17']:
    for esc_top_cell in row:
        esc_top_cell.border = top_border   
        
for row in sheetname['A19:H19']:
    for esc_bottom_cell in row:
        esc_bottom_cell.border = bottom_border 

for row in sheetname['A23:H23']:
    for fee_top_cell in row:
        fee_top_cell.border = top_border   
        
for row in sheetname['A26:H26']:
    for cost_top_cell in row:
        cost_top_cell.border = top_border

for row in sheetname['A29:H29']:
    for loc_top_cell in row:
        loc_top_cell.border = top_border 
        
for row in sheetname['A33:H33']:
    for chgoff_top_cell in row:
        chgoff_top_cell.border = top_border

for row in sheetname['A36:H36']:
    for unapp_top_cell in row:
        unapp_top_cell.border = top_border

for row in sheetname['A39:H39']:
    for inv_top_cell in row:
        inv_top_cell.border = top_border

for row in sheetname['A42:H42']:
    for part_top_cell in row:
        part_top_cell.border = top_border
 
for row in sheetname['A45:H45']:
    for ytdInt_top_cell in row:
        ytdInt_top_cell.border = top_border 

for row in sheetname['A48:H48']:
    for ytdLate_top_cell in row:
        ytdLate_top_cell.border = top_border 

        
#auto size the entire excel 
for idx, col in enumerate(sheetname.columns, 1):
    sheetname.column_dimensions[get_column_letter(idx)].auto_size = True




#tracking the last exception cell 
countCell = int(keys['startCell'])
print("Start creating exception worksheet tabs for balancing category with variances#####...." )

#Border the excel to medium 
for row in sheetname[f'A{countCell-1}:H{countCell-1}']:
    for expl_top_cell in row:
        expl_top_cell.border = top_border 
        expl_top_cell.fill = grey_fill_excptn_area 


list_with_value={}

#create worksheet for the exception 
for row in range(8, 55):
    Taskno=sheetname[f"G{row}"].value
    Taskname=sheetname[f"A{row}"].value
    if not Taskno ==None:
        Dmpdata_minus_dna_cell =(defaultvalue(sheetname[f"D{row}"].value), defaultvalue(sheetname[f"D{row+1}"].value), defaultvalue(sheetname[f"D{row+2}"].value),
                    - defaultvalue(sheetname[f"F{row}"].value), -defaultvalue(sheetname[f"F{row+1}"].value), - defaultvalue(sheetname[f"F{row+2}"].value)
                    )
        add_Dmpdata_minus_dna_cell = round(math.fsum(Dmpdata_minus_dna_cell),2)
        print(Taskname +" Variance# " + str(add_Dmpdata_minus_dna_cell))

        if not add_Dmpdata_minus_dna_cell ==0:
            mytable = str.maketrans("("," ")
            print(f"{Taskno.translate(mytable)}  {Taskname}")
            taskitems=f"{Taskno.translate(mytable)}  {Taskname}"
            exceptionsheet = wb.create_sheet(taskitems) 

            if Taskname.lower() == 'principal balance':
                #populating the exceptions    
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["PRINCIPALBALANCE"]=  countCell
                print(list_with_value)
                countCell+=1
                sheetname[f"A{countCell}"] = "    Loans with balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area 
                
            elif Taskname.lower() == 'accrued interest': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["ACCRUEDINTEREST"]=  countCell
                print(list_with_value)
                countCell+=1
                sheetname[f"A{countCell}"] = "    Loans with Int balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=4
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area 
                        
            elif Taskname.lower() == 'escrow balance': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["ESCROWBALANCE"]=  countCell
                print(list_with_value)
                countCell+=1
                sheetname[f"A{countCell}"] = "    Loans with Escrow balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=4
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
            
            elif Taskname.lower() == 'late charge': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["LATECHARGES"]= countCell
                print(list_with_value)
                countCell+=1
                sheetname[f"A{countCell}"] = "    Loans with Late Charge balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
                        
            elif Taskname.lower() == 'defferred fees': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["DEFFEREDFEES"]=  countCell
                print(list_with_value)
                countCell+=1
                sheetname[f"A{countCell}"] = "    Loans with Differred fees balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
                        
            elif Taskname.lower() == 'defferred costs': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["DEFFEREDCOSTS"]=  countCell
                print(list_with_value)
                countCell+=1
                sheetname[f"A{countCell}"] = "    Loans with Differred costs balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
                        
            elif Taskname.lower() == 'original comm (loc)': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["CREDITLIMIT"]=  countCell
                print(list_with_value)
                countCell+=1
                sheetname[f"A{countCell}"] = "    Loans with LOC balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=4
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
            
            elif Taskname.lower() == 'charge-off balance': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["CHARGEOFFBALANCE"]=  countCell
                print(list_with_value)
                countCell+=1
                sheetname[f"A{countCell}"] = "    Loans with charge-off balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=4
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
            
            elif Taskname.lower() == 'unapplied balance': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["UNAPPLIEDBALANCE"]= countCell
                print(list_with_value)
                countCell+=1
                sheetname[f"A{countCell}"] = "    Loans with Unapplied balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
             
            elif Taskname.lower() == 'investor balance': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["INVESTORBALANCE"]=  countCell
                print(list_with_value)
                countCell+=1
                sheetname[f"A{countCell}"] = "    Loans with Investor balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
                        
            elif Taskname.lower() == 'participation': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["PARTICIPATION"]=  countCell
                print(list_with_value)
                countCell+=1
                sheetname[f"A{countCell}"] = "    Loans with Participation balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
                        
            elif Taskname.lower() == 'ytd interest': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["YTDINTEREST"]=  countCell
                print(list_with_value)
                countCell+=1
                sheetname[f"A{countCell}"] = "    Loans with Ytd Interest balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
                        
            elif Taskname.lower() == 'ytd late charge paid': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["YTDLATECHARGEPAID"]=  countCell
                print(list_with_value)
                countCell+=1
                sheetname[f"A{countCell}"] = "    Loans with YTD late charge balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
                        
            elif Taskname.lower() == 'shadow accounting iptp': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["SHADOWACCOUNTINGIPTP"]=  countCell
                print(list_with_value)
                countCell+=1
                sheetname[f"A{countCell}"] = "    Loans with Shadow balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area




# Apply the grey fill to the specified range
for key, value in list_with_variance.items(): 
     changecell = str.maketrans("H","C")
     getcellnumber = str.maketrans("H"," ")
     cell=f"{value.translate(changecell)}" 
     cellno=f"{value.translate(getcellnumber)}" 
     empty_category = sheetname[cell].value
     if  empty_category ==None:
         end_cell = int(cellno.strip())+2
         grey_range = sheetname[f'A{cellno.strip()}:H{end_cell}']
         for row in grey_range:
              for cell in row:
                   cell.fill = grey_fill 
         



def updateException(index):
    return 'N' if index is None or index ==0 else 'Y'
    
   

try:
     # Establish the connection
    conn = pymysql.connect(host = keys['dbhost'],
    port = int(3306),
    user = keys['dbusername'],
    password = keys['dbpassword'],
    db = keys['dbname'])
except Exception as e:
     print(e)
     
# Create a cursor object
cursor = conn.cursor()
# Write and execute the SELECT query
queryvariances = keys["loanbalancingvariance"]
cursor.execute(queryvariances)

# Fetch all rows from the executed query
results = cursor.fetchall()

headers = [description[0] for description in cursor.description ]

print(headers)
      
       


# Print the results
for row in results:
    print("Exception to process -->", row[0])
    if row[0]=="PRINCIPALBALANCE":

        if updateException(row[1])=='Y':
            sheetname[f"F{list_with_value["PRINCIPALBALANCE"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             sheetname[f"A{list_with_value["PRINCIPALBALANCE"]+2}"] = row[2]
             sheetname[f"F{list_with_value["PRINCIPALBALANCE"]+2}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             sheetname[f"A{list_with_value["PRINCIPALBALANCE"]+3}"] = row[4]
             sheetname[f"F{list_with_value["PRINCIPALBALANCE"]+3}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             sheetname[f"A{list_with_value["PRINCIPALBALANCE"]+4}"] = row[6]
             sheetname[f"F{list_with_value["PRINCIPALBALANCE"]+4}"] = defaultvalue(row[7])
             
        #balance the exception 
        sheetname[f"H{list_with_value["PRINCIPALBALANCE"]+1}"] = F"= {list_with_variance["PRINCIPALBALANCE"]}-{f"F{list_with_value["PRINCIPALBALANCE"]+1}"}-{f"F{list_with_value["PRINCIPALBALANCE"]+2}"}-{f"F{list_with_value["PRINCIPALBALANCE"]+3}"}" 

        sheetname[f"H{list_with_value["PRINCIPALBALANCE"]+1}"].number_format = formattingAmount 
        
        
    elif row[0]=="ACCRUEDINTEREST":
        if updateException(row[1])=='Y':
            sheetname[f"F{list_with_value["ACCRUEDINTEREST"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             sheetname[f"A{list_with_value["ACCRUEDINTEREST"]+2}"] = row[2]
             sheetname[f"F{list_with_value["ACCRUEDINTEREST"]+2}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             sheetname[f"A{list_with_value["ACCRUEDINTEREST"]+3}"] = row[4]
             sheetname[f"F{list_with_value["ACCRUEDINTEREST"]+3}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             sheetname[f"A{list_with_value["ACCRUEDINTEREST"]+4}"] = row[6]
             sheetname[f"F{list_with_value["ACCRUEDINTEREST"]+4}"] = defaultvalue(row[7])
             
         #balance the exception 
        sheetname[f"H{list_with_value["ACCRUEDINTEREST"]+1}"] = F"= {list_with_variance["ACCRUEDINTEREST"]}-{f"F{list_with_value["ACCRUEDINTEREST"]+1}"}-{f"F{list_with_value["ACCRUEDINTEREST"]+2}"}-{f"F{list_with_value["ACCRUEDINTEREST"]+3}"}" 

        sheetname[f"H{list_with_value["ACCRUEDINTEREST"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="ESCROWBALANCE":
        if updateException(row[1])=='Y':
            sheetname[f"F{list_with_value["ESCROWBALANCE"]+1}"] = row[1]
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             sheetname[f"A{list_with_value["ESCROWBALANCE"]+2}"] = row[2]
             sheetname[f"F{list_with_value["ESCROWBALANCE"]+2}"] = row[3]
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             sheetname[f"A{list_with_value["ESCROWBALANCE"]+3}"] = row[4]
             sheetname[f"F{list_with_value["ESCROWBALANCE"]+3}"] = row[5]
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             sheetname[f"A{list_with_value["ESCROWBALANCE"]+4}"] = row[6]
             sheetname[f"F{list_with_value["ESCROWBALANCE"]+4}"] = row[7]
         #balance the exception 
        sheetname[f"H{list_with_value["ESCROWBALANCE"]+1}"] = F"= {list_with_variance["ESCROWBALANCE"]}-{f"F{list_with_value["ESCROWBALANCE"]+1}"}-{f"F{list_with_value["ESCROWBALANCE"]+2}"}-{f"F{list_with_value["ESCROWBALANCE"]+3}"}" 

        sheetname[f"H{list_with_value["ESCROWBALANCE"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="LATECHARGES":
        if updateException(row[1])=='Y':
            sheetname[f"F{list_with_value["LATECHARGES"]+1}"] = row[1]
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             sheetname[f"A{list_with_value["LATECHARGES"]+2}"] = row[2]
             sheetname[f"F{list_with_value["LATECHARGES"]+2}"] = row[3]
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             sheetname[f"A{list_with_value["LATECHARGES"]+3}"] = row[4]
             sheetname[f"F{list_with_value["LATECHARGES"]+3}"] = row[5]
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             sheetname[f"A{list_with_value["LATECHARGES"]+4}"] = row[6]
             sheetname[f"F{list_with_value["LATECHARGES"]+4}"] = row[7]
         #balance the exception 
        sheetname[f"H{list_with_value["LATECHARGES"]+1}"] = F"= {list_with_variance["LATECHARGES"]}-{f"F{list_with_value["LATECHARGES"]+1}"}-{f"F{list_with_value["LATECHARGES"]+2}"}-{f"F{list_with_value["LATECHARGES"]+3}"}" 

        sheetname[f"H{list_with_value["LATECHARGES"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="DEFFEREDFEES":
        if updateException(row[1])=='Y':
            sheetname[f"F{list_with_value["DEFFEREDFEES"]+1}"] = row[1]
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             sheetname[f"A{list_with_value["DEFFEREDFEES"]+2}"] = row[2]
             sheetname[f"F{list_with_value["DEFFEREDFEES"]+2}"] = row[3]
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             sheetname[f"A{list_with_value["DEFFEREDFEES"]+3}"] = row[4]
             sheetname[f"F{list_with_value["DEFFEREDFEES"]+3}"] = row[5]
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             sheetname[f"A{list_with_value["DEFFEREDFEES"]+4}"] = row[6]
             sheetname[f"F{list_with_value["DEFFEREDFEES"]+4}"] = row[7]
         #balance the exception 
        sheetname[f"H{list_with_value["DEFFEREDFEES"]+1}"] = F"= {list_with_variance["DEFFEREDFEES"]}-{f"F{list_with_value["DEFFEREDFEES"]+1}"}-{f"F{list_with_value["DEFFEREDFEES"]+2}"}-{f"F{list_with_value["DEFFEREDFEES"]+3}"}" 

        sheetname[f"H{list_with_value["DEFFEREDFEES"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="DEFFEREDCOSTS":
        if updateException(row[1])=='Y':
            sheetname[f"F{list_with_value["DEFFEREDCOSTS"]+1}"] = row[1]
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             sheetname[f"A{list_with_value["DEFFEREDCOSTS"]+2}"] = row[2]
             sheetname[f"F{list_with_value["DEFFEREDCOSTS"]+2}"] = row[3]
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             sheetname[f"A{list_with_value["DEFFEREDCOSTS"]+3}"] = row[4]
             sheetname[f"F{list_with_value["DEFFEREDCOSTS"]+3}"] = row[5]
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             sheetname[f"A{list_with_value["DEFFEREDCOSTS"]+4}"] = row[6]
             sheetname[f"F{list_with_value["DEFFEREDCOSTS"]+4}"] = row[7]
         #balance the exception 
        sheetname[f"H{list_with_value["DEFFEREDCOSTS"]+1}"] = F"= {list_with_variance["DEFFEREDCOSTS"]}-{f"F{list_with_value["DEFFEREDCOSTS"]+1}"}-{f"F{list_with_value["DEFFEREDCOSTS"]+2}"}-{f"F{list_with_value["DEFFEREDCOSTS"]+3}"}" 

        sheetname[f"H{list_with_value["DEFFEREDCOSTS"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="CREDITLIMIT":
        if updateException(row[1])=='Y':
            sheetname[f"F{list_with_value["CREDITLIMIT"]+1}"] = row[1]
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             sheetname[f"A{list_with_value["CREDITLIMIT"]+2}"] = row[2]
             sheetname[f"F{list_with_value["CREDITLIMIT"]+2}"] = row[3]
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             sheetname[f"A{list_with_value["CREDITLIMIT"]+3}"] = row[4]
             sheetname[f"F{list_with_value["CREDITLIMIT"]+3}"] = row[5]
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             sheetname[f"A{list_with_value["CREDITLIMIT"]+4}"] = row[6]
             sheetname[f"F{list_with_value["CREDITLIMIT"]+4}"] = row[7]
             
         #balance the exception 
        sheetname[f"H{list_with_value["CREDITLIMIT"]+1}"] = F"= {list_with_variance["CREDITLIMIT"]}-{f"F{list_with_value["CREDITLIMIT"]+1}"}-{f"F{list_with_value["CREDITLIMIT"]+2}"}-{f"F{list_with_value["CREDITLIMIT"]+3}"}" 

        sheetname[f"H{list_with_value["CREDITLIMIT"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="CHARGEOFFBALANCE":
        if updateException(row[1])=='Y':
            sheetname[f"F{list_with_value["CHARGEOFFBALANCE"]+1}"] = row[1]
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             sheetname[f"A{list_with_value["CHARGEOFFBALANCE"]+2}"] = row[2]
             sheetname[f"F{list_with_value["CHARGEOFFBALANCE"]+2}"] = row[3]
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             sheetname[f"A{list_with_value["CHARGEOFFBALANCE"]+3}"] = row[4]
             sheetname[f"F{list_with_value["CHARGEOFFBALANCE"]+3}"] = row[5]
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             sheetname[f"A{list_with_value["CHARGEOFFBALANCE"]+4}"] = row[6]
             sheetname[f"F{list_with_value["CHARGEOFFBALANCE"]+4}"] = row[7]
         #balance the exception 
        sheetname[f"H{list_with_value["CHARGEOFFBALANCE"]+1}"] = F"= {list_with_variance["CHARGEOFFBALANCE"]}-{f"F{list_with_value["CHARGEOFFBALANCE"]+1}"}-{f"F{list_with_value["CHARGEOFFBALANCE"]+2}"}-{f"F{list_with_value["CHARGEOFFBALANCE"]+3}"}" 

        sheetname[f"H{list_with_value["CHARGEOFFBALANCE"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="UNAPPLIEDBALANCE":
        if updateException(row[1])=='Y':
            sheetname[f"F{list_with_value["UNAPPLIEDBALANCE"]+1}"] = row[1]
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             sheetname[f"A{list_with_value["UNAPPLIEDBALANCE"]+2}"] = row[2]
             sheetname[f"F{list_with_value["UNAPPLIEDBALANCE"]+2}"] = row[3]
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             sheetname[f"A{list_with_value["UNAPPLIEDBALANCE"]+3}"] = row[4]
             sheetname[f"F{list_with_value["UNAPPLIEDBALANCE"]+3}"] = row[5]
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             sheetname[f"A{list_with_value["UNAPPLIEDBALANCE"]+4}"] = row[6]
             sheetname[f"F{list_with_value["UNAPPLIEDBALANCE"]+4}"] = row[7]
        #balance the exception 
        sheetname[f"H{list_with_value["UNAPPLIEDBALANCE"]+1}"] = F"= {list_with_variance["UNAPPLIEDBALANCE"]}-{f"F{list_with_value["UNAPPLIEDBALANCE"]+1}"}-{f"F{list_with_value["UNAPPLIEDBALANCE"]+2}"}-{f"F{list_with_value["UNAPPLIEDBALANCE"]+3}"}" 

        sheetname[f"H{list_with_value["UNAPPLIEDBALANCE"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="INVESTORBALANCE":
        if updateException(row[1])=='Y':
            sheetname[f"F{list_with_value["INVESTORBALANCE"]+1}"] = row[1]
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             sheetname[f"A{list_with_value["INVESTORBALANCE"]+2}"] = row[2]
             sheetname[f"F{list_with_value["INVESTORBALANCE"]+2}"] = row[3]
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             sheetname[f"A{list_with_value["INVESTORBALANCE"]+3}"] = row[4]
             sheetname[f"F{list_with_value["INVESTORBALANCE"]+3}"] = row[5]
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             sheetname[f"A{list_with_value["INVESTORBALANCE"]+4}"] = row[6]
             sheetname[f"F{list_with_value["INVESTORBALANCE"]+4}"] = row[7]
         #balance the exception 
        sheetname[f"H{list_with_value["INVESTORBALANCE"]+1}"] = F"= {list_with_variance["INVESTORBALANCE"]}-{f"F{list_with_value["INVESTORBALANCE"]+1}"}-{f"F{list_with_value["INVESTORBALANCE"]+2}"}-{f"F{list_with_value["INVESTORBALANCE"]+3}"}" 

        sheetname[f"H{list_with_value["INVESTORBALANCE"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="PARTICIPATION":
        if updateException(row[1])=='Y':
            sheetname[f"F{list_with_value["PARTICIPATION"]+1}"] = row[1]
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             sheetname[f"A{list_with_value["PARTICIPATION"]+2}"] = row[2]
             sheetname[f"F{list_with_value["PARTICIPATION"]+2}"] = row[3]
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             sheetname[f"A{list_with_value["PARTICIPATION"]+3}"] = row[4]
             sheetname[f"F{list_with_value["PARTICIPATION"]+3}"] = row[5]
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             sheetname[f"A{list_with_value["PARTICIPATION"]+4}"] = row[6]
             sheetname[f"F{list_with_value["PARTICIPATION"]+4}"] = row[7]
         #balance the exception 
        sheetname[f"H{list_with_value["PARTICIPATION"]+1}"] = F"= {list_with_variance["PARTICIPATION"]}-{f"F{list_with_value["PARTICIPATION"]+1}"}-{f"F{list_with_value["PARTICIPATION"]+2}"}-{f"F{list_with_value["PARTICIPATION"]+3}"}" 

        sheetname[f"H{list_with_value["PARTICIPATION"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="YTDINTEREST":
        if updateException(row[1])=='Y':
            sheetname[f"F{list_with_value["YTDINTEREST"]+1}"] = row[1]
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             sheetname[f"A{list_with_value["YTDINTEREST"]+2}"] = row[2]
             sheetname[f"F{list_with_value["YTDINTEREST"]+2}"] = row[3]
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             sheetname[f"A{list_with_value["YTDINTEREST"]+3}"] = row[4]
             sheetname[f"F{list_with_value["YTDINTEREST"]+3}"] = row[5]
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             sheetname[f"A{list_with_value["YTDINTEREST"]+4}"] = row[6]
             sheetname[f"F{list_with_value["YTDINTEREST"]+4}"] = row[7]
         #balance the exception 
        sheetname[f"H{list_with_value["YTDINTEREST"]+1}"] = F"= {list_with_variance["YTDINTEREST"]}-{f"F{list_with_value["YTDINTEREST"]+1}"}-{f"F{list_with_value["YTDINTEREST"]+2}"}-{f"F{list_with_value["YTDINTEREST"]+3}"}" 

        sheetname[f"H{list_with_value["YTDINTEREST"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="YTDLATECHARGEPAID":
        if updateException(row[1])=='Y':
            sheetname[f"F{list_with_value["YTDLATECHARGEPAID"]+1}"] = row[1]
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             sheetname[f"A{list_with_value["YTDLATECHARGEPAID"]+2}"] = row[2]
             sheetname[f"F{list_with_value["YTDLATECHARGEPAID"]+2}"] = row[3]
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             sheetname[f"A{list_with_value["YTDLATECHARGEPAID"]+3}"] = row[4]
             sheetname[f"F{list_with_value["YTDLATECHARGEPAID"]+3}"] = row[5]
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             sheetname[f"A{list_with_value["YTDLATECHARGEPAID"]+4}"] = row[6]
             sheetname[f"F{list_with_value["YTDLATECHARGEPAID"]+4}"] = row[7]
         #balance the exception 
        sheetname[f"H{list_with_value["YTDLATECHARGEPAID"]+1}"] = F"= {list_with_variance["YTDLATECHARGEPAID"]}-{f"F{list_with_value["YTDLATECHARGEPAID"]+1}"}-{f"F{list_with_value["YTDLATECHARGEPAID"]+2}"}-{f"F{list_with_value["YTDLATECHARGEPAID"]+3}"}" 

        sheetname[f"H{list_with_value["YTDLATECHARGEPAID"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="SHADOWACCOUNTINGIPTP":
        if updateException(row[1])=='Y':
            sheetname[f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+1}"] = row[1]
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             sheetname[f"A{list_with_value["SHADOWACCOUNTINGIPTP"]+2}"] = row[2]
             sheetname[f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+2}"] = row[3]
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             sheetname[f"A{list_with_value["SHADOWACCOUNTINGIPTP"]+3}"] = row[4]
             sheetname[f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+3}"] = row[5]
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             sheetname[f"A{list_with_value["SHADOWACCOUNTINGIPTP"]+4}"] = row[6]
             sheetname[f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+4}"] = row[7]
         #balance the exception 
        sheetname[f"H{list_with_value["SHADOWACCOUNTINGIPTP"]+1}"] = F"= {list_with_variance["SHADOWACCOUNTINGIPTP"]}-{f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+1}"}-{f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+2}"}-{f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+3}"}" 

        sheetname[f"H{list_with_value["SHADOWACCOUNTINGIPTP"]+1}"].number_format = formattingAmount 
   

cursor.close()
conn.close()


#format the currency amt for exception section
print("Number Format - exeception section")
for row in range(int(keys['startCell']),countCell):
     sheetname[f"F{row}"].number_format = formattingAmount  
 


alphabet =[
  'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
  'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ',
  'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ',
  'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ', 'CR', 'CS', 'CT', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ',
  'DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH', 'DI', 'DJ', 'DK', 'DL', 'DM', 'DN', 'DO', 'DP', 'DQ', 'DR', 'DS', 'DT', 'DU', 'DV', 'DW', 'DX', 'DY', 'DZ',
  'EA', 'EB', 'EC', 'ED', 'EE', 'EF', 'EG', 'EH', 'EI', 'EJ', 'EK', 'EL', 'EM', 'EN', 'EO', 'EP', 'EQ', 'ER', 'ES', 'ET', 'EU', 'EV', 'EW', 'EX', 'EY', 'EZ',
  'FA', 'FB', 'FC', 'FD', 'FE', 'FF', 'FG', 'FH', 'FI', 'FJ', 'FK', 'FL', 'FM', 'FN', 'FO', 'FP', 'FQ', 'FR', 'FS', 'FT', 'FU', 'FV', 'FW', 'FX', 'FY', 'FZ',
  'GA', 'GB', 'GC', 'GD', 'GE', 'GF', 'GG', 'GH', 'GI', 'GJ', 'GK', 'GL', 'GM', 'GN', 'GO', 'GP', 'GQ', 'GR', 'GS', 'GT', 'GU', 'GV', 'GW', 'GX', 'GY', 'GZ'
]


try:
     # Establish the connection
    conn = pymysql.connect(host = keys['dbhost'],
    port = int(3306),
    user = keys['dbusername'],
    password = keys['dbpassword'],
    db = keys['dbname'])
except Exception as  e:
     print(e)
     
# Create a cursor object
cursor = conn.cursor()

sheets = wb.sheetnames  
oldsheetmajor = ""
print("-->>> ", sheets) 
 

queryloop = "SELECT * FROM developer.tbl_query where reqyn='Y'"
cursor.execute(queryloop)
ListofException = cursor.fetchall()

for x in range(len(ListofException)):
     print("------------------------------------------------------------------------------------------------------------")
     oldsheetmajor_count=0
     
     data = ListofException[x] 

     retrieveQueryparameters = []  
     for i in data:
         retrieveQueryparameters.append(i)

     exceptQuery = retrieveQueryparameters[2] #Extract the query from database
     cursor.execute(exceptQuery)
     exceptionresults = cursor.fetchall()
    
     print(f"retrieveQueryparameters >>>>>>>  {retrieveQueryparameters} ")
     exceptionTitle = retrieveQueryparameters[3]
     print(f"retrieveQueryparameters >>>>>>>creating the ws title>>>  {exceptionTitle} ")

     #loop through exception worksheet to justify  
     for xcell in range(1,len(sheets)):
        key = sheets[xcell].strip() 

        match = re.search(retrieveQueryparameters[1], # major from tbl query
                            key.upper()) #exception sheet title 

        # MAJOR BALANCING WITH SIMILAR NAMES 
        matchsecondcheck = re.search('YTD LATE', # major from tbl query
                             retrieveQueryparameters[1]) #exception sheet title 

        
        if match:
                print(f">>>>match>>>>>  {match.group()} >>>>oldsheetmajor>>>>>> {oldsheetmajor}  >>> matchsecondcheck  >>>>   {matchsecondcheck}") 

                print("xcel>>>looping the ws titles "+str(xcell)) 

                # Get column names for the exception to populate worksheet
                columns = [desc[0] for desc in cursor.description]  

                #select a worksheet
                sheettab = wb[sheets[xcell]] # get the desired sheet name
                maxColumn=sheettab.max_column
                columnheader_count = len(columns)
               
                

                 # MAJOR BALANCING WITH SIMILAR NAMES ==> check late charge & ytd late charge 
                if sheettab.title.strip() == '13)  YTD Late Charge Paid' and matchsecondcheck==None:
                    print("------here---check late charge & ytd late charge ")
                    continue
                
                print(f"maxcolum of >>>>>>>> {sheettab.title} >>maxColumn>>  {maxColumn} columnheader count >>> {len(columns)} ")
        
                if maxColumn > 1:
                    oldsheetmajor_count=2
                    print("oldsheetmajor_count>>>" + str(oldsheetmajor_count))
 
                 #UPDATE WORKBOOK TITLES 
                if oldsheetmajor_count == 0:
                    sheettab.merge_cells("A1:Q1") 
                    sheettab['A1'] = exceptionTitle 
                    sheettab['A1'].alignment =alignment
                    sheettab['A1'].font =bold_font
                else:
                    print(f"aphabet>>>>  {alphabet[maxColumn]} ")
                    cell_start = maxColumn+2
                    cell_end=maxColumn+len(columns)
                    sheettab.merge_cells(f"{alphabet[cell_start]}1:{alphabet[cell_end]}1") 
                    sheettab[f'{alphabet[cell_start]}1'] = exceptionTitle 
                    sheettab[f'{alphabet[cell_start]}1'].alignment =alignment
                    sheettab[f'{alphabet[cell_start]}1'].font =bold_font

                    


                # Write column headers to the Excel sheet
                for col_num, column_title in enumerate(columns, start=1):
                    if maxColumn==1:
                        sheettab.cell(row=2, column=col_num, value=column_title)
                    else:
                        sheettab.cell(row=2, column=col_num+maxColumn+oldsheetmajor_count, value=column_title)


            # Write data rows to the Excel sheet
                for row_num, row_data in enumerate(exceptionresults, start=2):
                    for col_num, cell_value in enumerate(row_data, start=1):
                        if maxColumn==1:
                            sheettab.cell(row=row_num+1, column=col_num, value=cell_value)
                        else:
                            sheettab.cell(row=row_num+1, column=col_num+maxColumn+oldsheetmajor_count, value=cell_value)
            
cursor.close()
conn.close() 

wb.save(f"{keys["workbookFilename"]}.xlsx ")


print('Task completed successfully')


