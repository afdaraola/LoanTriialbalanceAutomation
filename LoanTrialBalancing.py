import sys
import math
import os
import openpyxl 
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font,PatternFill,Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import pymysql
#import pandas as pd
import mysql.connector
import re


#! python3

print('start calling up the excel')



#set the directory 
os.chdir("C:\\Users\\daram\\pythonscripts")

#display working directory
print(os.getcwd())

#extract parameters from property file 
seperator="=="
keys ={}

with open('config.txt') as f:
    for line in f:
        if seperator in line:
            name, value = line.split(seperator,1)
            keys[name.strip()] = value.strip()

print("parameters -> ", keys)


#load the workbook 
wb = openpyxl.load_workbook("Template.xlsx")

 # Center-align the content of the cell
alignment = Alignment(horizontal='center', vertical='center')

#greyout some rows 
# Define the grey fill style
grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="lightGray")

grey_fill_excptn_area = PatternFill(start_color="d4ebf2", end_color="d4ebf2", fill_type="solid")

#define default value
def defaultvalue(x):
    return 0 if x is None else x

#get the acctive sheet 
sheetname = wb["Cover Sheet"] 


# Define border styles
top_border = Border(
    top=Side(style='medium'),
    bottom=Side(style='thin')
    )

right_border = Border( 
    right=Side(style='medium')
    )

bottom_border = Border(
    top=Side(style='thin'),
    bottom=Side(style='medium')
    )
bottom_only_border = Border(
    bottom=Side(style='medium')
    )

# Define border styles
thick_border = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

# Define thin border styles
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
    )

#display working directory
print(os.getcwd())


formattingAmount = keys['formatt']
#define custant 
NETVAR= keys['variance']


# Create a comment
print("Add comment of accrued through date >>>>>")
comment_text = keys['commentText'] 
author = keys['authorComment']
comment = Comment(comment_text,author)

# Add the comment to the cell
sheetname['H4'].comment = comment



print("Number format parttern --> "+formattingAmount)

#set the headers 
#merge cells 
sheetname.merge_cells("A2:H2")
 #set title 
sheetname['A2'] = "Loan Conversion balance Sheet" 

# Define the range of cells to apply borders
cell_range = sheetname['A2:H2']

# Apply the border to each cell in the range
for row in sheetname['A7:H7']:
    for cell in row:
        cell.border = thin_border
        cell.fill = grey_fill

#fill the title 
for row in cell_range:
    for cell in row:
        cell.fill = grey_fill


# Center-align the content of the cell
sheetname['A2'].alignment = Alignment(horizontal='center', vertical='center')

# Apply bold formatting to the cell
bold_font = Font(bold=True)
sheetname['A2'].font = bold_font



 #set parameters 
sheetname['A3'] = "CLIENT NAME:"
sheetname['C3'] = keys['clientname']  
sheetname['E3'] = "CONVERSION/TAPE DATE:" 
sheetname['F3']=keys['conversiondate']
sheetname['E4'] = "ACCRUED THROUGH DATE:" 
sheetname['F4']=keys['accruedthrough']

 

sheetname['A5'] = "BALANCED BY:" 
sheetname['C5']=keys['balancedby']
sheetname['A6']= "MANAGER BRANCH:" 
sheetname['C6']=keys['branch']


sheetname['C6'].font = bold_font

sheetname['A7'] = "ITEM/TASK"  
sheetname['A7'].font = bold_font
sheetname['A7'].alignment = alignment
#Setting dmpdsts
#merge cells 
sheetname.merge_cells("C7:D7")
sheetname['C7'] = "CURRENT SYSTEM" 
sheetname['C7'].font = bold_font

sheetname['C7'].alignment = alignment


#Setting DNA
#merge cells 
sheetname.merge_cells("E7:F7")
sheetname['E7'] = "DNA SYSTEM" 
sheetname['E7'].font = bold_font
# Center-align the content of the cell
sheetname['E7'].alignment = alignment
 
 #Setting Variance
 #merge cells 
sheetname.merge_cells("G7:H7")
sheetname['G7'] = "VARIANCES- SEE EXPLANATIONS BELOW" 
sheetname['G7'].font = bold_font
# Center-align the content of the cell
sheetname['G7'].alignment = alignment


 
list_with_variance={}
tRow = 8
 
sheetname[f'A{tRow}'] = "Principal Balance"
sheetname[f'E{tRow}'] = "CM_TRIAL"
sheetname[f'E{tRow+1}'] = "CN_TRIAL"
sheetname[f'E{tRow+2}'] = "MG_TRIAL"
sheetname[f'G{tRow}'] = "(1)"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}"
#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['PRINCIPALBALANCE'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border   

tRow = tRow+3

sheetname[f'A{tRow}'] = "Accrued Interest"
sheetname[f'E{tRow}'] = "CM_TRIAL"
sheetname[f'E{tRow+1}'] = "CN_TRIAL"
sheetname[f'E{tRow+2}'] = "MG_TRIAL"
sheetname[f'G{tRow}'] = "(2)"

sheetname[f'A{tRow+3}'] = "Non-Accrued Interest"
sheetname[f'E{tRow+3}'] = "CM_TRIAL"
sheetname[f'E{tRow+4}'] = "CN_TRIAL"
sheetname[f'E{tRow+5}'] = "MG_TRIAL"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}-F{tRow+3}-f{tRow+4}-f{tRow+5}"
#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['ACCRUEDINTEREST'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border   

tRow = tRow+6

sheetname[f'A{tRow}'] = "Escrow Balance"
sheetname[f'E{tRow}'] = "CM_TRIAL"
sheetname[f'E{tRow+1}'] = "CN_TRIAL"
sheetname[f'E{tRow+2}'] = "MG_TRIAL"
sheetname[f'G{tRow}'] = "(3)"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}"
#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['ESCROWBALANCE'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+3

sheetname[f'A{tRow}'] = "Loans In Process (LIP)"
sheetname[f'E{tRow}'] = "CM_TRIAL"
sheetname[f'E{tRow+1}'] = "CN_TRIAL"
sheetname[f'E{tRow+2}'] = "MG_TRIAL"
sheetname[f'G{tRow}'] =  "(4)"
sheetname[f'H{tRow}'] =  f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}"
#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['LOANSINPROCESS'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+3

sheetname[f'A{tRow}'] = "Late Charge"
sheetname[f'E{tRow}'] = "CM_TRIAL"
sheetname[f'E{tRow+1}'] = "CN_TRIAL"
sheetname[f'E{tRow+2}'] = "MG_TRIAL"
sheetname[f'G{tRow}'] = "(5)"
sheetname[f'H{tRow}'] =  f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}"

#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['LATECHARGES'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+3

sheetname[f'A{tRow}'] = "Deferred Fees"
sheetname[f'E{tRow}'] = "LN_AMRT" 
sheetname[f'G{tRow}'] = "(6)"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}" 

#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['DEFERREDFEES'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+3

sheetname[f'A{tRow}'] = "Deferred Costs"
sheetname[f'E{tRow}'] = "LN_AMRT" 
sheetname[f'G{tRow}'] = "(7)"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}" 
#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['DEFERREDCOSTS'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+3

sheetname[f'A{tRow}'] = "Original Comm (LOC)"
sheetname[f'E{tRow}'] = "CM_TRIAL"
sheetname[f'E{tRow+1}'] = "CN_TRIAL"
sheetname[f'E{tRow+2}'] = "MG_TRIAL"
sheetname[f'E{tRow+3}'] = "ML_TRIAL"
sheetname[f'G{tRow}'] = "(8)"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}-F{tRow+3}"

#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['CREDITLIMIT'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+4

sheetname[f'A{tRow}'] = "Charge-off Balance"
sheetname[f'E{tRow}'] = "CM_TRIAL"
sheetname[f'E{tRow+1}'] = "CN_TRIAL"
sheetname[f'E{tRow+2}'] = "MG_TRIAL"
sheetname[f'G{tRow}'] = "(9)"
sheetname[f'H{tRow}'] =  f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}"
#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['CHARGEOFFBALANCE'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+3

sheetname[f'A{tRow}'] = "Unapplied Balance"
sheetname[f'E{tRow}'] = "CM_TRIAL"
sheetname[f'E{tRow+1}'] = "CN_TRIAL"
sheetname[f'E{tRow+2}'] = "MG_TRIAL"
sheetname[f'G{tRow}'] = "(10)"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}" 
#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['UNAPPLIEDBALANCE'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+3

sheetname[f'A{tRow}'] = "Investor Balance"
sheetname[f'E{tRow}'] = "CM_TRIAL/IL_TRIAL"
sheetname[f'E{tRow+1}'] = "CN_TRIAL/IL_TRIAL"
sheetname[f'E{tRow+2}'] = "MG_TRIAL/IL_TRIAL"
sheetname[f'G{tRow}'] = "(11)"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}" 
#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['INVESTORBALANCE'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+3

sheetname[f'A{tRow}'] = "Participation"
sheetname[f'E{tRow}'] = "CM_TRIAL"
sheetname[f'E{tRow+1}'] = "CN_TRIAL"
sheetname[f'E{tRow+2}'] = "MG_TRIAL"
sheetname[f'G{tRow}'] = "(12)"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}" 

#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['PARTICIPATION'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+3

sheetname[f'A{tRow}'] = "YTD Interest"
sheetname[f'E{tRow}'] = "CM_TRIAL"
sheetname[f'E{tRow+1}'] = "CN_TRIAL"
sheetname[f'E{tRow+2}'] = "MG_TRIAL"
sheetname[f'G{tRow}'] = "(13)"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}" 

#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['YTDINTEREST'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+3

sheetname[f'A{tRow}'] = "YTD Late Charge Paid"
sheetname[f'E{tRow}'] = "MM_YTD"
sheetname[f'G{tRow}'] = "(14)"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}" 

#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['YTDLATECHARGEPAID'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+3

sheetname[f'A{tRow}'] = "YTD Taxes"
sheetname[f'E{tRow}'] = "CM_TRIAL"
sheetname[f'E{tRow+1}'] = "CN_TRIAL"
sheetname[f'E{tRow+2}'] = "MG_TRIAL"
sheetname[f'G{tRow}'] = "(15)"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}" 

#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['YTDTAXES'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+3

sheetname[f'A{tRow}'] = "YTD Points Paid"
sheetname[f'E{tRow}'] = "MM_YTD"
sheetname[f'G{tRow}'] = "(16)"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}" 

#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['YTDPOINTSPAID'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+3

sheetname[f'A{tRow}'] = "YTD PMI Paid"
sheetname[f'E{tRow}'] = "MM_YTD"
sheetname[f'G{tRow}'] = "(17)"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}" 

#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['YTDPMIPAID'] = f'H{tRow}'


#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+3

sheetname[f'A{tRow}'] = "Shadow Accounting IPTP"
sheetname[f'E{tRow}'] = "CM_TRIAL"
sheetname[f'E{tRow+1}'] = "CN_TRIAL"
sheetname[f'E{tRow+2}'] = "MG_TRIAL"
sheetname[f'G{tRow}'] = "(18)"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}" 

#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['SHADOWACCOUNTINGIPTP'] = f'H{tRow}'

#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

tRow = tRow+3

sheetname[f'A{tRow}'] = "Forebearance Balance"
sheetname[f'E{tRow}'] = "MG_TRIAL"
sheetname[f'E{tRow+1}'] = "CN_TRIAL"
sheetname[f'E{tRow+2}'] = "MG_TRIAL"
sheetname[f'G{tRow}'] = "(19)"
sheetname[f'H{tRow}'] = f"=D{tRow}+D{tRow+1}+D{tRow+2}-F{tRow}-F{tRow+1}-F{tRow+2}" 

#format currency
sheetname[f'H{tRow}'].number_format = formattingAmount 
list_with_variance['FOREBEARANCEBALANCE'] = f'H{tRow}'


#border the line 
for row in sheetname[f'A{tRow}:H{tRow}']:
    for prin_top_cell in row:
        prin_top_cell.border = top_border  

#border the line 
for row in sheetname[f'A{tRow+2}:H{tRow+2}']:
    for prin_top_cell in row:
        prin_top_cell.border = bottom_border  

#Connection to DB 
print("Connection to DB.....")
try:
     # Establish the connection
    conn = pymysql.connect(host = keys['dbhost'],
    port = int(3306),
    user = keys['dbusername'],
    password = keys['dbpassword'],
    db = keys['dbname'])
except Exception as e:
     print(e)
# Check if the connection is successful
if conn:
     print ("connection successful") 

# Create a cursor object
cursor = conn.cursor()
 
# Write and execute the SELECT query
query = keys["loanbalancing"] #query in property file 
cursor.execute(query)

# Fetch all rows from the executed query
results = cursor.fetchall()

print(f"list_with_variance >>>>, {list_with_variance} list_with_variance['PRINCIPALBALANCE'] >> {list_with_variance['PRINCIPALBALANCE'][1:]}")

lRow = 0
dmpAmtypp=0
dnaSumypp=0
dmpAmtyppts=0
dnaSumyppts=0
dmpAmtyplate=0
dnaSumyplate=0
dnaSumFee=0 
dmpAmtFee=0 
dmpAmtCost=0 
dnaSumCost=0
# Print the results
for row in results:
    #print(row)
    
    if row[1] =='PRINCIPALBALANCE':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['PRINCIPALBALANCE'][1:])
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2]
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2]
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2]
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5]
    elif row[1] =='ACCRUEDINTEREST':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['ACCRUEDINTEREST'][1:])
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2]
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2]
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2]
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
    elif row[1] =='ESCROWBALANCE':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['ESCROWBALANCE'][1:])
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2]
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2]
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2]
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
    elif row[1] =='LOANSINPROCESS':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['LOANSINPROCESS'][1:])
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2]
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2]
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2]
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
    elif row[1] =='LATECHARGES':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['LATECHARGES'][1:])
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2]
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2]
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2]
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
    elif row[1] =='DEFERREDFEES':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['DEFERREDFEES'][1:]) 
        if row[4]=='CML' or row[4]=='CNS' or row[4]=='MTG':
            if row[2] is not None:
                sheetname[f'C{lRow}'] = row[2]
            if  row[3] is not None:
                dmpAmtFee = dmpAmtFee + row[3]
                sheetname[f'D{lRow}'] = dmpAmtFee
            if row[5] is not None:
                dnaSumFee = dnaSumFee + row[5]
                sheetname[f'F{lRow}'] = dnaSumFee
        """
        if row[4]=='CML': 
            sheetname[f'C{lRow}'] = row[2]
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2]
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2]
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
        """
    elif row[1] =='DEFERREDCOSTS':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['DEFERREDCOSTS'][1:])
        if row[4]=='CML' or row[4]=='CNS' or row[4]=='MTG':
            if row[2] is not None:
                sheetname[f'C{lRow}'] = row[2]
            if  row[3] is not None:
                dmpAmtCost = dmpAmtCost + row[3]
                sheetname[f'D{lRow}'] = dmpAmtCost
            if row[5] is not None:
                dnaSumCost = dnaSumCost + row[5]
                sheetname[f'F{lRow}'] = dnaSumCost
        """
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2]
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2]
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2]
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
            """
    elif row[1] =='CREDITLIMIT':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['CREDITLIMIT'][1:])
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2]
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2]
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2]
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5]
        if row[4]=='MLN':
            sheetname[f'C{lRow+3}'] = row[2]
            sheetname[f'D{lRow+3}'] = row[3]
            sheetname[f'F{lRow+3}'] = row[5] 
    elif row[1] =='CHARGEOFFBALANCE':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['CHARGEOFFBALANCE'][1:])
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2]
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2]
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2]
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
    elif row[1] =='UNAPPLIEDBALANCE':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['UNAPPLIEDBALANCE'][1:])
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2]
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2]
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2]
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
    elif row[1] =='INVESTORBALANCE':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['INVESTORBALANCE'][1:])
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2]
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2]
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2]
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
    elif row[1] =='PARTICIPATION':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['PARTICIPATION'][1:])
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2]
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2]
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2]
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
    elif row[1] =='YTDINTEREST':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['YTDINTEREST'][1:])
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2]
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2]
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2]
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
    elif row[1] =='YTDLATECHARGEPAID':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['YTDLATECHARGEPAID'][1:])
        if row[4]=='CML' or row[4]=='CNS' or row[4]=='MTG':
            if row[2] is not None:
                sheetname[f'C{lRow}'] = row[2]
            if  row[3] is not None:
                dmpAmtyplate = dmpAmtyplate + row[3]
                sheetname[f'D{lRow}'] = dmpAmtyplate
            if row[5] is not None:
                dnaSumyplate = dnaSumyplate + row[5]
                sheetname[f'F{lRow}'] = dnaSumyplate
        """
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2] 
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2] 
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2] 
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
            """
    elif row[1] =='YTDTAXES':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['YTDTAXES'][1:])
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2] 
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2] 
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2] 
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
    elif row[1] =='YTDPOINTSPAID':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['YTDPOINTSPAID'][1:])
        if row[4]=='CML' or row[4]=='CNS' or row[4]=='MTG':
            if row[2] is not None:
                sheetname[f'C{lRow}'] = row[2]
            if  row[3] is not None:
                dmpAmtyppts = dmpAmtyppts + row[3]
                sheetname[f'D{lRow}'] = dmpAmtyppts
            if row[5] is not None:
                dnaSumyppts = dnaSumyppts + row[5]
                sheetname[f'F{lRow}'] = dnaSumyppts
        """
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2] 
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2] 
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2] 
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
            """
    elif row[1] =='YTDPMIPAID':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['YTDPMIPAID'][1:])
        if row[4]=='CML' or row[4]=='CNS' or row[4]=='MTG':
            if row[2] is not None:
                sheetname[f'C{lRow}'] = row[2]
            if  row[3] is not None:
                dmpAmtypp = dmpAmtypp + row[3]
                sheetname[f'D{lRow}'] = dmpAmtypp
            if row[5] is not None:
                dnaSumypp = dnaSumypp + row[5]
                sheetname[f'F{lRow}'] = dnaSumypp
        """
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2] 
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2] 
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2] 
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
            """
    elif row[1] =='SHADOWACCOUNTINGIPTP':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['SHADOWACCOUNTINGIPTP'][1:])
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2]
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2]
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2]
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
    elif row[1] =='FOREBEARANCEBALANCE':
        print("BAL CATEGORY ->>", row[1] , row[4])
        lRow = int(list_with_variance['FOREBEARANCEBALANCE'][1:])
        if row[4]=='CML':
            sheetname[f'C{lRow}'] = row[2]
            sheetname[f'D{lRow}'] = row[3]
            sheetname[f'F{lRow}'] = row[5]
        if row[4]=='CNS':
            sheetname[f'C{lRow+1}'] = row[2]
            sheetname[f'D{lRow+1}'] = row[3]
            sheetname[f'F{lRow+1}'] = row[5]
        if row[4]=='MTG':
            sheetname[f'C{lRow+2}'] = row[2]
            sheetname[f'D{lRow+2}'] = row[3]
            sheetname[f'F{lRow+2}'] = row[5] 
    

else:
    print ("Clossing connection")
# Close the connection when done 
cursor.close()
conn.close()    

print(" Max rows >>>>>>>>>>>>> I am here ",sheetname.max_row)
#bold row E
for row in range(8, sheetname.max_row+2):
    cell = sheetname[f"E{row}"]
   # cell.font =bold_font
    cell.alignment = alignment
    #format currency amount DNA values
    currencycellDNA = sheetname[f"F{row}"] 
    currencycellDNA.number_format = formattingAmount 
    #format currency amount Dmpdata values
    currencycellDmp = sheetname[f"D{row}"] 
    currencycellDmp.number_format = formattingAmount 

    dmpdatabalsource = sheetname[f"C{row}"]
    dmpdatabalsource.alignment = alignment

    balanceNumber = sheetname[f"G{row}"]
    balanceNumber.alignment = alignment
    
    

# Add themed text to the column
color_font = Font(color="FF0000")

        
#auto size the entire excel 
for idx, col in enumerate(sheetname.columns, 1):
    sheetname.column_dimensions[get_column_letter(idx)].auto_size = True




#tracking the last exception cell 
countCell = int(keys['startCell'])

print("Start creating exception worksheet tabs for balancing category with variances#####...." )

sheetname[f'A{countCell-2}'] = "Explanation of any out of balance conditions:" 

#Border the excel to medium 
for row in sheetname[f'A{countCell}:H{countCell}']:
    for expl_top_cell in row:
        expl_top_cell.border = top_border 
        expl_top_cell.fill = grey_fill_excptn_area 


list_with_value={}

#create worksheet for the exception 
for row in range(8, countCell-4):
    Taskno=sheetname[f"G{row}"].value
    Taskname=sheetname[f"A{row}"].value
    if not Taskno ==None:
        Dmpdata_minus_dna_cell =(defaultvalue(sheetname[f"D{row}"].value), defaultvalue(sheetname[f"D{row+1}"].value), defaultvalue(sheetname[f"D{row+2}"].value),
                    - defaultvalue(sheetname[f"F{row}"].value), -defaultvalue(sheetname[f"F{row+1}"].value), - defaultvalue(sheetname[f"F{row+2}"].value)
                    )
        add_Dmpdata_minus_dna_cell = round(math.fsum(Dmpdata_minus_dna_cell),2)
        print(Taskname +" Variance# " + str(add_Dmpdata_minus_dna_cell))

        if not add_Dmpdata_minus_dna_cell ==0:
            mytable = str.maketrans("("," ")
            print(f"{Taskno.translate(mytable)}  {Taskname}")
            taskitems=f"{Taskno.translate(mytable)}  {Taskname}"
            exceptionsheet = wb.create_sheet(taskitems) 

            if Taskname.lower() == 'principal balance':
                #populating the exceptions    
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["PRINCIPALBALANCE"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area 
                
            elif Taskname.lower() == 'accrued interest': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["ACCRUEDINTEREST"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with Int balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=4
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area 
                        
            elif Taskname.lower() == 'escrow balance': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["ESCROWBALANCE"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with Escrow balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=4
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
            
            elif Taskname.lower() == 'late charge': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["LATECHARGES"]= countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with Late Charge balances not converting" #chnages to exceptn  
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
            
            elif Taskname.lower() == 'loans in process (lip)': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["LOANSINPROCESS"]= countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with LIP balances not converting" #chnages to exceptn  
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
                        
            elif Taskname.lower() == 'deferred fees': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["DEFERREDFEES"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with Differred fees balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
                        
            elif Taskname.lower() == 'deferred costs': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["DEFERREDCOSTS"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with Differred costs balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
                        
            elif Taskname.lower() == 'original comm (loc)': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["CREDITLIMIT"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with LOC balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=4
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
            
            elif Taskname.lower() == 'charge-off balance': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["CHARGEOFFBALANCE"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with charge-off balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=4
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
            
            elif Taskname.lower() == 'unapplied balance': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["UNAPPLIEDBALANCE"]= countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with Unapplied balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
             
            elif Taskname.lower() == 'investor balance': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["INVESTORBALANCE"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with Investor balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
                        
            elif Taskname.lower() == 'participation': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["PARTICIPATION"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with Participation balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
                        
            elif Taskname.lower() == 'ytd interest': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["YTDINTEREST"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with Ytd Interest balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
                        
            elif Taskname.lower() == 'ytd late charge paid': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["YTDLATECHARGEPAID"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with YTD late charge balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
            
            elif Taskname.lower() == 'ytd taxes': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["YTDTAXES"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with YTD late charge balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
            
            elif Taskname.lower() == 'ytd points paid': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["YTDPOINTSPAID"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with YTD late charge balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
            
            elif Taskname.lower() == 'ytd pmi paid': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["YTDPMIPAID"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with YTD late charge balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area
                        

                        
            elif Taskname.lower() == 'shadow accounting iptp': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["SHADOWACCOUNTINGIPTP"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with Shadow balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area

            elif Taskname.lower() == 'forebearance balance': 
                sheetname[f"A{countCell}"] = taskitems
                list_with_value["FOREBEARANCEBALANCE"]=  countCell
                print(list_with_value)
                countCell+=1
                #sheetname[f"A{countCell}"] = "    Loans with YTD late charge balances not converting" #chnages to exceptn
                sheetname[f"G{countCell}"] = NETVAR
                sheetname[f"G{countCell}"].font  = color_font
                sheetname[f"H{countCell}"] = add_Dmpdata_minus_dna_cell
                countCell+=3
                 #Create grey fill 
                for row in sheetname[f"A{countCell}": f"H{countCell}"]:
                    for expl_top_cell in row:
                        expl_top_cell.border = top_border 
                        expl_top_cell.fill = grey_fill_excptn_area



# Apply the grey fill to the specified range
for key, value in list_with_variance.items(): 
     changecell = str.maketrans("H","C")
     getcellnumber = str.maketrans("H"," ")
     cell=f"{value.translate(changecell)}" 
     cellno=f"{value.translate(getcellnumber)}" 
     splitcellAlp = cell[:1]
     splitcellnum = cell[1:]
     increasecell= f"{splitcellAlp}{int(splitcellnum)+1}"
     increasecell2 = f"{splitcellAlp}{int(splitcellnum)+2}"
     empty_category = sheetname[cell].value
     #print(f"Festus grey increasecell>>> {increasecell} increasecell2 >>>> {increasecell2} cell >>>>>> {cell} cellno>>>>> {cellno} value>>>>{value} ")
     empty_category1 = sheetname[increasecell].value
     empty_category2 = sheetname[increasecell2].value
     if  empty_category ==None and empty_category1 ==None and empty_category2 ==None:
         end_cell = int(cellno.strip())+2
         grey_range = sheetname[f'A{cellno.strip()}:H{end_cell}']
         for row in grey_range:
              for cell in row:
                   cell.fill = grey_fill 
         



def updateException(index):
    return 'N' if index is None or index ==0 else 'Y'
    
   

try:
     # Establish the connection
    conn = pymysql.connect(host = keys['dbhost'],
    port = int(3306),
    user = keys['dbusername'],
    password = keys['dbpassword'],
    db = keys['dbname'])
except Exception as e:
     print(e)
     
# Create a cursor object
cursor = conn.cursor()
# Write and execute the SELECT query
queryvariances = keys["loanbalancingvariance"]
cursor.execute(queryvariances)

# Fetch all rows from the executed query
results = cursor.fetchall()

headers = [description[0] for description in cursor.description ]

print(headers)
      
       
list_of_exception_position =[]

# Print the results
for row in results:
    print("Exception to process -->", row[0])
    if row[0]=="PRINCIPALBALANCE":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["PRINCIPALBALANCE"]+1}"] =  "Loans with balances not converting"
            sheetname[f"F{list_with_value["PRINCIPALBALANCE"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
            if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["PRINCIPALBALANCE"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["PRINCIPALBALANCE"]+2}"] = defaultvalue(row[3])
            else:
                sheetname[f"A{list_with_value["PRINCIPALBALANCE"]+1}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["PRINCIPALBALANCE"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
            if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["PRINCIPALBALANCE"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["PRINCIPALBALANCE"]+3}"] = defaultvalue(row[5])
            else:
                sheetname[f"A{list_with_value["PRINCIPALBALANCE"]+2}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["PRINCIPALBALANCE"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["PRINCIPALBALANCE"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["PRINCIPALBALANCE"]+4}"] = defaultvalue(row[7])
             else:
                sheetname[f"A{list_with_value["PRINCIPALBALANCE"]+3}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["PRINCIPALBALANCE"]+3}"] = defaultvalue(row[7])
             
        #balance the exception 
        sheetname[f"H{list_with_value["PRINCIPALBALANCE"]+1}"] = F"= {list_with_variance["PRINCIPALBALANCE"]}-{f"F{list_with_value["PRINCIPALBALANCE"]+1}"}-{f"F{list_with_value["PRINCIPALBALANCE"]+2}"}-{f"F{list_with_value["PRINCIPALBALANCE"]+3}"}" 

        sheetname[f"H{list_with_value["PRINCIPALBALANCE"]+1}"].number_format = formattingAmount 
        
        #print("list_of_exception_position ==========", list_of_exception_position)
        
    elif row[0]=="ACCRUEDINTEREST":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["ACCRUEDINTEREST"]+1}"] = "Loans with Int balances not converting"
            sheetname[f"F{list_with_value["ACCRUEDINTEREST"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["ACCRUEDINTEREST"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["ACCRUEDINTEREST"]+2}"] = defaultvalue(row[3])
             else:
                 sheetname[f"A{list_with_value["ACCRUEDINTEREST"]+1}"] = row[2].capitalize()
                 sheetname[f"F{list_with_value["ACCRUEDINTEREST"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["ACCRUEDINTEREST"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["ACCRUEDINTEREST"]+3}"] = defaultvalue(row[5])
             else:
                 sheetname[f"A{list_with_value["ACCRUEDINTEREST"]+2}"] = row[4].capitalize()
                 sheetname[f"F{list_with_value["ACCRUEDINTEREST"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["ACCRUEDINTEREST"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["ACCRUEDINTEREST"]+4}"] = defaultvalue(row[7])
             else:
                 sheetname[f"A{list_with_value["ACCRUEDINTEREST"]+3}"] = row[6].capitalize()
                 sheetname[f"F{list_with_value["ACCRUEDINTEREST"]+3}"] = defaultvalue(row[7])
             
         #balance the exception 
        sheetname[f"H{list_with_value["ACCRUEDINTEREST"]+1}"] = F"= {list_with_variance["ACCRUEDINTEREST"]}-{f"F{list_with_value["ACCRUEDINTEREST"]+1}"}-{f"F{list_with_value["ACCRUEDINTEREST"]+2}"}-{f"F{list_with_value["ACCRUEDINTEREST"]+3}"}" 

        sheetname[f"H{list_with_value["ACCRUEDINTEREST"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="ESCROWBALANCE":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["ESCROWBALANCE"]+1}"] = "Loans with Escrow balances not converting"
            sheetname[f"F{list_with_value["ESCROWBALANCE"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["ESCROWBALANCE"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["ESCROWBALANCE"]+2}"] = defaultvalue(row[3])
             else:
                 sheetname[f"A{list_with_value["ESCROWBALANCE"]+1}"] = row[2].capitalize()
                 sheetname[f"F{list_with_value["ESCROWBALANCE"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
            if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["ESCROWBALANCE"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["ESCROWBALANCE"]+3}"] = defaultvalue(row[5])
            else:
                sheetname[f"A{list_with_value["ESCROWBALANCE"]+2}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["ESCROWBALANCE"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
            if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["ESCROWBALANCE"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["ESCROWBALANCE"]+4}"] = defaultvalue(row[7])
            else:
                sheetname[f"A{list_with_value["ESCROWBALANCE"]+3}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["ESCROWBALANCE"]+3}"] = defaultvalue(row[7])
         #balance the exception 
        sheetname[f"H{list_with_value["ESCROWBALANCE"]+1}"] = F"= {list_with_variance["ESCROWBALANCE"]}-{f"F{list_with_value["ESCROWBALANCE"]+1}"}-{f"F{list_with_value["ESCROWBALANCE"]+2}"}-{f"F{list_with_value["ESCROWBALANCE"]+3}"}" 

        sheetname[f"H{list_with_value["ESCROWBALANCE"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="LATECHARGES":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["LATECHARGES"]+1}"] = "Loans with Late Charge balances not converting"
            sheetname[f"F{list_with_value["LATECHARGES"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["LATECHARGES"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["LATECHARGES"]+2}"] = defaultvalue(row[3])
             else:
                 sheetname[f"A{list_with_value["LATECHARGES"]+1}"] = row[2].capitalize()
                 sheetname[f"F{list_with_value["LATECHARGES"]+1}"] = defaultvalue(row[3])
                 
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["LATECHARGES"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["LATECHARGES"]+3}"] = defaultvalue(row[5])
             else:
                 sheetname[f"A{list_with_value["LATECHARGES"]+2}"] = row[4].capitalize()
                 sheetname[f"F{list_with_value["LATECHARGES"]+2}"] = defaultvalue(row[5])
                 
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["LATECHARGES"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["LATECHARGES"]+4}"] = defaultvalue(row[7])
             else:
                 sheetname[f"A{list_with_value["LATECHARGES"]+3}"] = row[6].capitalize()
                 sheetname[f"F{list_with_value["LATECHARGES"]+3}"] = defaultvalue(row[7])
         #balance the exception 
        sheetname[f"H{list_with_value["LATECHARGES"]+1}"] = F"= {list_with_variance["LATECHARGES"]}-{f"F{list_with_value["LATECHARGES"]+1}"}-{f"F{list_with_value["LATECHARGES"]+2}"}-{f"F{list_with_value["LATECHARGES"]+3}"}" 

        sheetname[f"H{list_with_value["LATECHARGES"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="LOANSINPROCESS":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["LOANSINPROCESS"]+1}"] = "Loans with LIP balances not converting"
            sheetname[f"F{list_with_value["LOANSINPROCESS"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
              if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["LOANSINPROCESS"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["LOANSINPROCESS"]+2}"] = defaultvalue(row[3])
              else:
                  sheetname[f"A{list_with_value["LOANSINPROCESS"]+1}"] = row[2].capitalize()
                  sheetname[f"F{list_with_value["LOANSINPROCESS"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                 sheetname[f"A{list_with_value["LOANSINPROCESS"]+3}"] = row[4].capitalize()
                 sheetname[f"F{list_with_value["LOANSINPROCESS"]+3}"] = defaultvalue(row[5])
             else:
                 sheetname[f"A{list_with_value["LOANSINPROCESS"]+2}"] = row[4].capitalize()
                 sheetname[f"F{list_with_value["LOANSINPROCESS"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
              if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                 sheetname[f"A{list_with_value["LOANSINPROCESS"]+4}"] = row[6].capitalize()
                 sheetname[f"F{list_with_value["LOANSINPROCESS"]+4}"] = defaultvalue(row[7])
              else:
                 sheetname[f"A{list_with_value["LOANSINPROCESS"]+3}"] = row[6].capitalize()
                 sheetname[f"F{list_with_value["LOANSINPROCESS"]+3}"] = defaultvalue(row[7])
         #balance the exception 
        sheetname[f"H{list_with_value["LOANSINPROCESS"]+1}"] = F"= {list_with_variance["LOANSINPROCESS"]}-{f"F{list_with_value["LOANSINPROCESS"]+1}"}-{f"F{list_with_value["LOANSINPROCESS"]+2}"}-{f"F{list_with_value["LOANSINPROCESS"]+3}"}" 

        sheetname[f"H{list_with_value["LOANSINPROCESS"]+1}"].number_format = formattingAmount 

    elif row[0]=="DEFERREDFEES":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["DEFERREDFEES"]+1}"] = "Loans with Differred fees balances not converting"
            sheetname[f"F{list_with_value["DEFERREDFEES"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["DEFERREDFEES"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["DEFERREDFEES"]+2}"] = defaultvalue(row[3])
             else:
                sheetname[f"A{list_with_value["DEFERREDFEES"]+1}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["DEFERREDFEES"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["DEFERREDFEES"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["DEFERREDFEES"]+3}"] = defaultvalue(row[5])
             else:
                sheetname[f"A{list_with_value["DEFERREDFEES"]+2}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["DEFERREDFEES"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["DEFERREDFEES"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["DEFERREDFEES"]+4}"] = defaultvalue(row[7])
             else:
                sheetname[f"A{list_with_value["DEFERREDFEES"]+3}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["DEFERREDFEES"]+3}"] = defaultvalue(row[7])
         #balance the exception 
        sheetname[f"H{list_with_value["DEFERREDFEES"]+1}"] = F"= {list_with_variance["DEFERREDFEES"]}-{f"F{list_with_value["DEFERREDFEES"]+1}"}-{f"F{list_with_value["DEFERREDFEES"]+2}"}-{f"F{list_with_value["DEFERREDFEES"]+3}"}" 

        sheetname[f"H{list_with_value["DEFERREDFEES"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="DEFERREDCOSTS":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["DEFERREDCOSTS"]+1}"] = "Loans with Differred Costs balances not converting"
            sheetname[f"F{list_with_value["DEFERREDCOSTS"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["DEFERREDCOSTS"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["DEFERREDCOSTS"]+2}"] = defaultvalue(row[3])
             else:
                sheetname[f"A{list_with_value["DEFERREDCOSTS"]+1}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["DEFERREDCOSTS"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["DEFERREDCOSTS"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["DEFERREDCOSTS"]+3}"] = defaultvalue(row[5])
             else:
                sheetname[f"A{list_with_value["DEFERREDCOSTS"]+2}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["DEFERREDCOSTS"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
              if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["DEFERREDCOSTS"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["DEFERREDCOSTS"]+4}"] = defaultvalue(row[7])
              else:
                sheetname[f"A{list_with_value["DEFERREDCOSTS"]+3}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["DEFERREDCOSTS"]+3}"] = defaultvalue(row[7])
         #balance the exception 
        sheetname[f"H{list_with_value["DEFERREDCOSTS"]+1}"] = F"= {list_with_variance["DEFERREDCOSTS"]}-{f"F{list_with_value["DEFERREDCOSTS"]+1}"}-{f"F{list_with_value["DEFERREDCOSTS"]+2}"}-{f"F{list_with_value["DEFERREDCOSTS"]+3}"}" 

        sheetname[f"H{list_with_value["DEFERREDCOSTS"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="CREDITLIMIT":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["CREDITLIMIT"]+1}"] = "Loans with Credit Limit balances not converting"
            sheetname[f"F{list_with_value["CREDITLIMIT"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["CREDITLIMIT"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["CREDITLIMIT"]+2}"] = defaultvalue(row[3])
             else:
                sheetname[f"A{list_with_value["CREDITLIMIT"]+1}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["CREDITLIMIT"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
              if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["CREDITLIMIT"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["CREDITLIMIT"]+3}"] = defaultvalue(row[5])
              else:
                sheetname[f"A{list_with_value["CREDITLIMIT"]+2}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["CREDITLIMIT"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["CREDITLIMIT"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["CREDITLIMIT"]+4}"] = defaultvalue(row[7])
             else:
                sheetname[f"A{list_with_value["CREDITLIMIT"]+3}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["CREDITLIMIT"]+3}"] = defaultvalue(row[7])
             
         #balance the exception 
        sheetname[f"H{list_with_value["CREDITLIMIT"]+1}"] = F"= {list_with_variance["CREDITLIMIT"]}-{f"F{list_with_value["CREDITLIMIT"]+1}"}-{f"F{list_with_value["CREDITLIMIT"]+2}"}-{f"F{list_with_value["CREDITLIMIT"]+3}"}" 

        sheetname[f"H{list_with_value["CREDITLIMIT"]+1}"].number_format = formattingAmount 
        
       # print("list_of_exception_position ====CREDITLIMIT======", list_of_exception_position)

    elif row[0]=="CHARGEOFFBALANCE":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["CHARGEOFFBALANCE"]+1}"] ="Loans with Charge-off balances not converting"
            sheetname[f"F{list_with_value["CHARGEOFFBALANCE"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["CHARGEOFFBALANCE"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["CHARGEOFFBALANCE"]+2}"] = defaultvalue(row[3])
             else:
                sheetname[f"A{list_with_value["CHARGEOFFBALANCE"]+1}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["CHARGEOFFBALANCE"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["CHARGEOFFBALANCE"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["CHARGEOFFBALANCE"]+3}"] = defaultvalue(row[5])
             else:
                sheetname[f"A{list_with_value["CHARGEOFFBALANCE"]+2}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["CHARGEOFFBALANCE"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
              if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["CHARGEOFFBALANCE"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["CHARGEOFFBALANCE"]+4}"] = defaultvalue(row[7])
              else:
                sheetname[f"A{list_with_value["CHARGEOFFBALANCE"]+3}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["CHARGEOFFBALANCE"]+3}"] = defaultvalue(row[7])
         #balance the exception 
        sheetname[f"H{list_with_value["CHARGEOFFBALANCE"]+1}"] = F"= {list_with_variance["CHARGEOFFBALANCE"]}-{f"F{list_with_value["CHARGEOFFBALANCE"]+1}"}-{f"F{list_with_value["CHARGEOFFBALANCE"]+2}"}-{f"F{list_with_value["CHARGEOFFBALANCE"]+3}"}" 

        sheetname[f"H{list_with_value["CHARGEOFFBALANCE"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="UNAPPLIEDBALANCE":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["UNAPPLIEDBALANCE"]+1}"] = "Loans with Unapplied balances not converting"
            sheetname[f"F{list_with_value["UNAPPLIEDBALANCE"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["UNAPPLIEDBALANCE"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["UNAPPLIEDBALANCE"]+2}"] = defaultvalue(row[3])
             else:
                sheetname[f"A{list_with_value["UNAPPLIEDBALANCE"]+1}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["UNAPPLIEDBALANCE"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
              if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["UNAPPLIEDBALANCE"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["UNAPPLIEDBALANCE"]+3}"] = defaultvalue(row[5])
              else:
                sheetname[f"A{list_with_value["UNAPPLIEDBALANCE"]+2}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["UNAPPLIEDBALANCE"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["UNAPPLIEDBALANCE"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["UNAPPLIEDBALANCE"]+4}"] = defaultvalue(row[7])
             else:
                sheetname[f"A{list_with_value["UNAPPLIEDBALANCE"]+3}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["UNAPPLIEDBALANCE"]+3}"] = defaultvalue(row[7])
        #balance the exception 
        sheetname[f"H{list_with_value["UNAPPLIEDBALANCE"]+1}"] = F"= {list_with_variance["UNAPPLIEDBALANCE"]}-{f"F{list_with_value["UNAPPLIEDBALANCE"]+1}"}-{f"F{list_with_value["UNAPPLIEDBALANCE"]+2}"}-{f"F{list_with_value["UNAPPLIEDBALANCE"]+3}"}" 

        sheetname[f"H{list_with_value["UNAPPLIEDBALANCE"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="INVESTORBALANCE":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["INVESTORBALANCE"]+1}"] = "Loans with Investor balances not converting"
            sheetname[f"F{list_with_value["INVESTORBALANCE"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["INVESTORBALANCE"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["INVESTORBALANCE"]+2}"] = defaultvalue(row[3])
             else:
                sheetname[f"A{list_with_value["INVESTORBALANCE"]+1}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["INVESTORBALANCE"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["INVESTORBALANCE"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["INVESTORBALANCE"]+3}"] = defaultvalue(row[5])
             else:
                sheetname[f"A{list_with_value["INVESTORBALANCE"]+2}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["INVESTORBALANCE"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["INVESTORBALANCE"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["INVESTORBALANCE"]+4}"] = defaultvalue(row[7])
             else:
                sheetname[f"A{list_with_value["INVESTORBALANCE"]+3}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["INVESTORBALANCE"]+3}"] = defaultvalue(row[7])
         #balance the exception 
        sheetname[f"H{list_with_value["INVESTORBALANCE"]+1}"] = F"= {list_with_variance["INVESTORBALANCE"]}-{f"F{list_with_value["INVESTORBALANCE"]+1}"}-{f"F{list_with_value["INVESTORBALANCE"]+2}"}-{f"F{list_with_value["INVESTORBALANCE"]+3}"}" 

        sheetname[f"H{list_with_value["INVESTORBALANCE"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="PARTICIPATION":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["PARTICIPATION"]+1}"] = "Loans with Participation balances not converting"
            sheetname[f"F{list_with_value["PARTICIPATION"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["PARTICIPATION"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["PARTICIPATION"]+2}"] = defaultvalue(row[3])
             else:
                sheetname[f"A{list_with_value["PARTICIPATION"]+1}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["PARTICIPATION"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["PARTICIPATION"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["PARTICIPATION"]+3}"] = defaultvalue(row[5])
             else:
                sheetname[f"A{list_with_value["PARTICIPATION"]+2}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["PARTICIPATION"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
              if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["PARTICIPATION"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["PARTICIPATION"]+4}"] = defaultvalue(row[7])
              else:
                sheetname[f"A{list_with_value["PARTICIPATION"]+3}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["PARTICIPATION"]+3}"] = defaultvalue(row[7])
         #balance the exception 
        sheetname[f"H{list_with_value["PARTICIPATION"]+1}"] = F"= {list_with_variance["PARTICIPATION"]}-{f"F{list_with_value["PARTICIPATION"]+1}"}-{f"F{list_with_value["PARTICIPATION"]+2}"}-{f"F{list_with_value["PARTICIPATION"]+3}"}" 

        sheetname[f"H{list_with_value["PARTICIPATION"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="YTDINTEREST":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["YTDINTEREST"]+1}"] = "Loans with YTD interest balances not converting"
            sheetname[f"F{list_with_value["YTDINTEREST"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
             if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["YTDINTEREST"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["YTDINTEREST"]+2}"] = defaultvalue(row[3])
             else:
                sheetname[f"A{list_with_value["YTDINTEREST"]+1}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["YTDINTEREST"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
             if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["YTDINTEREST"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["YTDINTEREST"]+3}"] = defaultvalue(row[5])
             else:
                sheetname[f"A{list_with_value["YTDINTEREST"]+2}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["YTDINTEREST"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
             if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["YTDINTEREST"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["YTDINTEREST"]+4}"] = defaultvalue(row[7])
             else:
                sheetname[f"A{list_with_value["YTDINTEREST"]+3}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["YTDINTEREST"]+3}"] = defaultvalue(row[7])
         #balance the exception 
        sheetname[f"H{list_with_value["YTDINTEREST"]+1}"] = F"= {list_with_variance["YTDINTEREST"]}-{f"F{list_with_value["YTDINTEREST"]+1}"}-{f"F{list_with_value["YTDINTEREST"]+2}"}-{f"F{list_with_value["YTDINTEREST"]+3}"}" 

        sheetname[f"H{list_with_value["YTDINTEREST"]+1}"].number_format = formattingAmount 
        
    elif row[0]=="YTDLATECHARGEPAID":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["YTDLATECHARGEPAID"]+1}"] = "Loans with YTD late charge unpaid balances not converting"
            sheetname[f"F{list_with_value["YTDLATECHARGEPAID"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
            if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["YTDLATECHARGEPAID"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["YTDLATECHARGEPAID"]+2}"] = defaultvalue(row[3])
            else:
                sheetname[f"A{list_with_value["YTDLATECHARGEPAID"]+1}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["YTDLATECHARGEPAID"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
            if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["YTDLATECHARGEPAID"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["YTDLATECHARGEPAID"]+3}"] = defaultvalue(row[5])
            else:
                sheetname[f"A{list_with_value["YTDLATECHARGEPAID"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["YTDLATECHARGEPAID"]+3}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
            if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["YTDLATECHARGEPAID"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["YTDLATECHARGEPAID"]+4}"] = defaultvalue(row[7])
            else:
                sheetname[f"A{list_with_value["YTDLATECHARGEPAID"]+3}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["YTDLATECHARGEPAID"]+3}"] = defaultvalue(row[7])
         #balance the exception 
        sheetname[f"H{list_with_value["YTDLATECHARGEPAID"]+1}"] = F"= {list_with_variance["YTDLATECHARGEPAID"]}-{f"F{list_with_value["YTDLATECHARGEPAID"]+1}"}-{f"F{list_with_value["YTDLATECHARGEPAID"]+2}"}-{f"F{list_with_value["YTDLATECHARGEPAID"]+3}"}" 

        sheetname[f"H{list_with_value["YTDLATECHARGEPAID"]+1}"].number_format = formattingAmount 
    
    elif row[0]=="YTDTAXES":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["YTDTAXES"]+1}"] = "Loans with YTD Taxes balances not converting"
            sheetname[f"F{list_with_value["YTDTAXES"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
            if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["YTDTAXES"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["YTDTAXES"]+2}"] = defaultvalue(row[3])
            else:
                sheetname[f"A{list_with_value["YTDTAXES"]+1}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["YTDTAXES"]+1}"] = defaultvalue(row[3]) 
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
            if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["YTDTAXES"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["YTDTAXES"]+3}"] = defaultvalue(row[5])
            else:
                sheetname[f"A{list_with_value["YTDTAXES"]+2}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["YTDTAXES"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
            if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["YTDTAXES"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["YTDTAXES"]+4}"] = defaultvalue(row[7])
            else:
                sheetname[f"A{list_with_value["YTDTAXES"]+3}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["YTDTAXES"]+3}"] = defaultvalue(row[7])
         #balance the exception 
        sheetname[f"H{list_with_value["YTDTAXES"]+1}"] = F"= {list_with_variance["YTDTAXES"]}-{f"F{list_with_value["YTDTAXES"]+1}"}-{f"F{list_with_value["YTDTAXES"]+2}"}-{f"F{list_with_value["YTDTAXES"]+3}"}" 

        sheetname[f"H{list_with_value["YTDTAXES"]+1}"].number_format = formattingAmount 

    elif row[0]=="YTDPOINTSPAID":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["YTDPOINTSPAID"]+1}"] = "Loans with YTD points paid balances not converting"
            sheetname[f"F{list_with_value["YTDPOINTSPAID"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
            if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["YTDPOINTSPAID"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["YTDPOINTSPAID"]+2}"] = defaultvalue(row[3])
            else:
                sheetname[f"A{list_with_value["YTDPOINTSPAID"]+1}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["YTDPOINTSPAID"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
            if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["YTDPOINTSPAID"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["YTDPOINTSPAID"]+3}"] = defaultvalue(row[5])
            else:
                sheetname[f"A{list_with_value["YTDPOINTSPAID"]+2}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["YTDPOINTSPAID"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
            if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["YTDPOINTSPAID"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["YTDPOINTSPAID"]+4}"] = defaultvalue(row[7])
            else:
                sheetname[f"A{list_with_value["YTDPOINTSPAID"]+3}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["YTDPOINTSPAID"]+3}"] = defaultvalue(row[7])
         #balance the exception 
        sheetname[f"H{list_with_value["YTDPOINTSPAID"]+1}"] = F"= {list_with_variance["YTDPOINTSPAID"]}-{f"F{list_with_value["YTDPOINTSPAID"]+1}"}-{f"F{list_with_value["YTDPOINTSPAID"]+2}"}-{f"F{list_with_value["YTDPOINTSPAID"]+3}"}" 

        sheetname[f"H{list_with_value["YTDPOINTSPAID"]+1}"].number_format = formattingAmount 

    elif row[0]=="YTDPMIPAID":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["YTDPMIPAID"]+1}"] = "Loans with YTD PMI paid balances not converting"
            sheetname[f"F{list_with_value["YTDPMIPAID"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
            if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["YTDPMIPAID"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["YTDPMIPAID"]+2}"] = defaultvalue(row[3])
            else:
                sheetname[f"A{list_with_value["YTDPMIPAID"]+1}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["YTDPMIPAID"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
            if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["YTDPMIPAID"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["YTDPMIPAID"]+3}"] = defaultvalue(row[5])
            else:
                sheetname[f"A{list_with_value["YTDPMIPAID"]+2}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["YTDPMIPAID"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
            if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["YTDPMIPAID"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["YTDPMIPAID"]+4}"] = defaultvalue(row[7])
            else:
                sheetname[f"A{list_with_value["YTDPMIPAID"]+3}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["YTDPMIPAID"]+3}"] = defaultvalue(row[7])
         #balance the exception 
        sheetname[f"H{list_with_value["YTDPMIPAID"]+1}"] = F"= {list_with_variance["YTDPMIPAID"]}-{f"F{list_with_value["YTDPMIPAID"]+1}"}-{f"F{list_with_value["YTDPMIPAID"]+2}"}-{f"F{list_with_value["YTDPMIPAID"]+3}"}" 

        sheetname[f"H{list_with_value["YTDPMIPAID"]+1}"].number_format = formattingAmount 

    elif row[0]=="SHADOWACCOUNTINGIPTP":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["SHADOWACCOUNTINGIPTP"]+1}"] = "Loans with Shadow balances not converting"
            sheetname[f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
            if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["SHADOWACCOUNTINGIPTP"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+2}"] = defaultvalue(row[3])
            else:
                sheetname[f"A{list_with_value["SHADOWACCOUNTINGIPTP"]+1}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+1}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
            if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["SHADOWACCOUNTINGIPTP"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+3}"] = defaultvalue(row[5])
            else:
                sheetname[f"A{list_with_value["SHADOWACCOUNTINGIPTP"]+2}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
            if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["SHADOWACCOUNTINGIPTP"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+4}"] = defaultvalue(row[7])
            else:
                sheetname[f"A{list_with_value["SHADOWACCOUNTINGIPTP"]+3}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+3}"] = defaultvalue(row[7])
         #balance the exception 
        sheetname[f"H{list_with_value["SHADOWACCOUNTINGIPTP"]+1}"] = F"= {list_with_variance["SHADOWACCOUNTINGIPTP"]}-{f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+1}"}-{f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+2}"}-{f"F{list_with_value["SHADOWACCOUNTINGIPTP"]+3}"}" 

        sheetname[f"H{list_with_value["SHADOWACCOUNTINGIPTP"]+1}"].number_format = formattingAmount 
   
    elif row[0]=="FOREBEARANCEBALANCE":
        if updateException(row[1])=='Y':
            sheetname[f"A{list_with_value["FOREBEARANCEBALANCE"]+1}"] = "Loans with YTD forebearance balances not converting"
            sheetname[f"F{list_with_value["FOREBEARANCEBALANCE"]+1}"] = defaultvalue(row[1])
        if updateException(row[2])=='Y' and updateException(row[3])=='Y':
            if updateException(row[1])=='Y':
                sheetname[f"A{list_with_value["FOREBEARANCEBALANCE"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["FOREBEARANCEBALANCE"]+2}"] = defaultvalue(row[3])
            else:
                sheetname[f"A{list_with_value["FOREBEARANCEBALANCE"]+2}"] = row[2].capitalize()
                sheetname[f"F{list_with_value["FOREBEARANCEBALANCE"]+2}"] = defaultvalue(row[3])
        if updateException(row[4])=='Y' and updateException(row[5])=='Y':
            if updateException(row[2])=='Y' and updateException(row[3])=='Y':
                sheetname[f"A{list_with_value["FOREBEARANCEBALANCE"]+3}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["FOREBEARANCEBALANCE"]+3}"] = defaultvalue(row[5])
            else:
                sheetname[f"A{list_with_value["FOREBEARANCEBALANCE"]+2}"] = row[4].capitalize()
                sheetname[f"F{list_with_value["FOREBEARANCEBALANCE"]+2}"] = defaultvalue(row[5])
        if updateException(row[6])=='Y' and updateException(row[7])=='Y':
            if updateException(row[4])=='Y' and updateException(row[5])=='Y':
                sheetname[f"A{list_with_value["FOREBEARANCEBALANCE"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["FOREBEARANCEBALANCE"]+4}"] = defaultvalue(row[7])
            else:
                sheetname[f"A{list_with_value["FOREBEARANCEBALANCE"]+4}"] = row[6].capitalize()
                sheetname[f"F{list_with_value["FOREBEARANCEBALANCE"]+4}"] = defaultvalue(row[7])
         #balance the exception 
        sheetname[f"H{list_with_value["FOREBEARANCEBALANCE"]+1}"] = F"= {list_with_variance["FOREBEARANCEBALANCE"]}-{f"F{list_with_value["FOREBEARANCEBALANCE"]+1}"}-{f"F{list_with_value["FOREBEARANCEBALANCE"]+2}"}-{f"F{list_with_value["FOREBEARANCEBALANCE"]+3}"}" 

        sheetname[f"H{list_with_value["FOREBEARANCEBALANCE"]+1}"].number_format = formattingAmount 
        

cursor.close()
conn.close()


#format the currency amt for exception section
print("Number Format - exeception section")
for row in range(int(keys['startCell']),countCell):
     sheetname[f"F{row}"].number_format = formattingAmount  
 


alphabet =[
  'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
  'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ',
  'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ',
  'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ', 'CR', 'CS', 'CT', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ',
  'DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH', 'DI', 'DJ', 'DK', 'DL', 'DM', 'DN', 'DO', 'DP', 'DQ', 'DR', 'DS', 'DT', 'DU', 'DV', 'DW', 'DX', 'DY', 'DZ',
  'EA', 'EB', 'EC', 'ED', 'EE', 'EF', 'EG', 'EH', 'EI', 'EJ', 'EK', 'EL', 'EM', 'EN', 'EO', 'EP', 'EQ', 'ER', 'ES', 'ET', 'EU', 'EV', 'EW', 'EX', 'EY', 'EZ',
  'FA', 'FB', 'FC', 'FD', 'FE', 'FF', 'FG', 'FH', 'FI', 'FJ', 'FK', 'FL', 'FM', 'FN', 'FO', 'FP', 'FQ', 'FR', 'FS', 'FT', 'FU', 'FV', 'FW', 'FX', 'FY', 'FZ',
  'GA', 'GB', 'GC', 'GD', 'GE', 'GF', 'GG', 'GH', 'GI', 'GJ', 'GK', 'GL', 'GM', 'GN', 'GO', 'GP', 'GQ', 'GR', 'GS', 'GT', 'GU', 'GV', 'GW', 'GX', 'GY', 'GZ'
]


try:
     # Establish the connection
    conn = pymysql.connect(host = keys['dbhost'],
    port = int(3306),
    user = keys['dbusername'],
    password = keys['dbpassword'],
    db = keys['dbname'])
except Exception as  e:
     print(e)
     
# Create a cursor object
cursor = conn.cursor()

sheets = wb.sheetnames  
oldsheetmajor = ""
print("-->>> ", sheets)  
 

queryloop = keys['loanqueryExceptions']
cursor.execute(queryloop)
ListofException = cursor.fetchall()

for x in range(len(ListofException)):
     print("------------------------------------------------------------------------------------------------------------")
     oldsheetmajor_count=0
     
     data = ListofException[x] 

     retrieveQueryparameters = []  
     for i in data:
         retrieveQueryparameters.append(i)

     exceptQuery = retrieveQueryparameters[2] #Extract the query from database
     cursor.execute(exceptQuery)
     exceptionresults = cursor.fetchall()
    
     print(f"retrieveQueryparameters >>>>>>>  {retrieveQueryparameters} ")
     exceptionTitle = retrieveQueryparameters[3]
     sumcolumnNo = retrieveQueryparameters[5] -1
     print(f"retrieveQueryparameters >>>>>>>creating the ws title>>>  {exceptionTitle}  >>>>sumcolumnNo >>> {sumcolumnNo}")

     #loop through exception worksheet to justify  
     for xcell in range(1,len(sheets)):
        key = sheets[xcell].strip() 
        print(f"key ==>>>{key}")

        match = re.search(retrieveQueryparameters[1], # major from tbl query
                            key.upper()) #exception sheet title 

        # MAJOR BALANCING WITH SIMILAR NAMES 
        matchsecondcheck = re.search('YTD LATE', # major from tbl query
                             retrieveQueryparameters[1]) #exception sheet title 

        #print(f">>>>match>>>>>  {match}  matchsecondcheck >>>>>{matchsecondcheck}")
        
        if match:
                print(f">>>>match>>>>>  {match.group()} >>>>oldsheetmajor>>>>>> {oldsheetmajor}  >>> matchsecondcheck  >>>>   {matchsecondcheck}") 

                print("xcel>>>looping the ws titles "+str(xcell)) 

                # Get column names for the exception to populate worksheet
                columns = [desc[0] for desc in cursor.description]  

                #select a worksheet
                sheettab = wb[sheets[xcell]] # get the desired sheet name
                maxColumn=sheettab.max_column
                columnheader_count = len(columns)
                
               
                

                 # MAJOR BALANCING WITH SIMILAR NAMES ==> check late charge & ytd late charge 
                if sheettab.title.strip() == '13)  YTD Late Charge Paid' and matchsecondcheck==None:
                    print("------here---check late charge & ytd late charge ")
                    continue
                
                print(f"maxcolum of >>>>>>>> {sheettab.title} >>maxColumn>>  {maxColumn} columnheader count >>> {len(columns)} >>> alphabet[sumcolumnNo-1] >>> {alphabet[sumcolumnNo-1]}")
        
                if maxColumn > 1:
                    oldsheetmajor_count=2
                    print("oldsheetmajor_count>>>" + str(oldsheetmajor_count))

                 #UPDATE WORKBOOK TITLES 
                if oldsheetmajor_count == 0:
                    cell_start = maxColumn-1
                    #cell_end=maxColumn+len(columns) 
                    cell_end=len(columns)-1 
                    sheettab.merge_cells(f"{alphabet[cell_start]}1:{alphabet[cell_end]}1") 
                    sheettab['A1'] = exceptionTitle 
                    sheettab['A1'].alignment =alignment
                    sheettab['A1'].font =bold_font
                    #placeholder 
                    placeholderColumnToSum = alphabet[sumcolumnNo]
                else:
                    print(f"aphabet>>>>  {alphabet[maxColumn]} ")
                    #cell_start = maxColumn+1
                    cell_start = maxColumn
                    cell_end=maxColumn+len(columns)
                    sheettab.merge_cells(f"{alphabet[cell_start]}1:{alphabet[cell_end-1]}1") 
                    sheettab[f'{alphabet[cell_start]}1'] = exceptionTitle 
                    sheettab[f'{alphabet[cell_start]}1'].alignment =alignment
                    sheettab[f'{alphabet[cell_start]}1'].font =bold_font
                    #placeholder 
                    placeholderColumnToSum = alphabet[cell_start+sumcolumnNo]

                    


                # Write column headers to the Excel sheet
                for col_num, column_title in enumerate(columns, start=1):
                    if maxColumn==1:
                        sheettab.cell(row=2, column=col_num, value=column_title)
                    else:
                        #sheettab.cell(row=2, column=col_num+maxColumn+oldsheetmajor_count, value=column_title)
                        sheettab.cell(row=2, column=col_num+maxColumn, value=column_title)


            # Write data rows to the Excel sheet
                for row_num, row_data in enumerate(exceptionresults, start=2):
                    for col_num, cell_value in enumerate(row_data, start=1):
                        if maxColumn==1:
                            sheettab.cell(row=row_num+1, column=col_num, value=cell_value)
                        else:
                            #sheettab.cell(row=row_num+1, column=col_num+maxColumn+oldsheetmajor_count, value=cell_value)
                            sheettab.cell(row=row_num+1, column=col_num+maxColumn, value=cell_value)
                
             #sum the column to get exception 
                maxrow = sheettab.max_row 
                print(f"placeholderColumnToSum >>> {placeholderColumnToSum} >>>>> maxrow>>>> {maxrow}  >>>alphabet[cell_end+1]>>>>>> {alphabet[cell_end+1]}")
                if maxColumn==1:
                    sheettab[f'{alphabet[cell_end+1]}2'] = f"=sum({placeholderColumnToSum}3:{placeholderColumnToSum}{maxrow})"
                    sheettab[f'{alphabet[cell_end+1]}2'].font=color_font
                    sheettab[f'{alphabet[cell_end+1]}2'].number_format = formattingAmount  
                else:
                    sheettab[f'{alphabet[cell_end]}2'] = f"=sum({placeholderColumnToSum}3:{placeholderColumnToSum}{maxrow})"
                    sheettab[f'{alphabet[cell_end]}2'].font=color_font
                    sheettab[f'{alphabet[cell_end]}2'].number_format = formattingAmount 
            
cursor.close()
conn.close() 


sheettab =  wb["Cover Sheet"]

signTab =  sheettab.max_row + 4 

print(f"signTab >>>>{signTab} ")

sheettab[f"A{signTab}"] = "SIGNED:" 

for row in sheetname[f'B{signTab}:D{signTab}']:
    for expl_top_cell in row:
        expl_top_cell.border = bottom_only_border  

sheettab[f"E{signTab}"] = "DATE:"

for row in sheetname[f'F{signTab}:H{signTab}']:
    for expl_top_cell in row:
        expl_top_cell.border = bottom_only_border  

wb.save(f"{keys["workbookFilename"]}.xlsx ")


print('Task completed successfully')


